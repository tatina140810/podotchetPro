"""READ-ONLY снимок финансовых агрегатов прода (сверка ДО/ПОСЛЕ деплоя soft-delete).

Бьёт по ЖИВОМУ серверу (127.0.0.1:8001) через stdlib urllib — без TestClient/httpx,
чтобы не ставить пакеты в прод-venv. Сервер отражает задеплоенный код (ДО = старый,
ПОСЛЕ = новый). Только GET + read-only POST /find-duplicates. Токены минтятся напрямую
(auth.create_access_token) для director-level юзера каждой организации.

  cd /root/PodotchetPRO/backend && .venv/bin/python prod_fingerprint.py /root/fp_before.json

Совпадение sha256 ДО и ПОСЛЕ (на данных без удалённых) = хук ничего не сломал.
"""
import hashlib
import io
import json
import sys
import urllib.error
import urllib.parse
import urllib.request

from openpyxl import load_workbook

import auth
import database
from models import Organization, User

BASE = "http://127.0.0.1:8001"
MONTHS = [(2026, m) for m in range(1, 13)]
FROM, TO = "2026-01-01", "2026-12-31"
DIR_ROLES = ("superadmin", "gen_director", "admin")


def canon(v):
    if isinstance(v, (tuple, list)):
        return sorted((canon(x) for x in v), key=lambda x: json.dumps(x, sort_keys=True, ensure_ascii=False, default=str))
    if isinstance(v, dict):
        return {k: canon(v[k]) for k in sorted(v)}
    return v


def _next_month(y, m):
    return (y + 1, 1) if m == 12 else (y, m + 1)


def full_expenses(tok):
    """Полный список расходов БЕЗ усечения (детерминированно): бьём по месяцам
    (каждый чанк < лимита) + бакет «до 2026». Устраняет недетерминизм LIMIT-200 +
    равные spent_at на границе. Дедуп по id, сортировка по id."""
    seen = {}
    buckets = [("", "2026-01-01")]  # всё до 2026
    for (y, mo) in MONTHS:
        ny, nm = _next_month(y, mo)
        buckets.append((f"{y}-{mo:02d}-01", f"{ny}-{nm:02d}-01"))
    for frm, to in buckets:
        params = {"limit": 1000, "date_to": to}
        if frm:
            params["date_from"] = frm
        rows = _req("GET", "/api/expenses", tok, params)
        if isinstance(rows, list) and rows and isinstance(rows[0], dict):
            for e in rows:
                seen[e["id"]] = e
    return [seen[i] for i in sorted(seen)]


def _req(method, path, token, params=None, body=None, raw=False):
    url = BASE + path
    if params:
        url += "?" + urllib.parse.urlencode(params)
    data = json.dumps(body).encode() if body is not None else None
    req = urllib.request.Request(url, data=data, method=method)
    req.add_header("Authorization", f"Bearer {token}")
    if body is not None:
        req.add_header("Content-Type", "application/json")
    try:
        with urllib.request.urlopen(req, timeout=60) as resp:
            content = resp.read()
        return content if raw else json.loads(content)
    except urllib.error.HTTPError as e:
        return ["err", e.code]


def xlsx_cells(content):
    if isinstance(content, list):  # ошибка
        return content
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    out = {ws.title: [list(r) for r in ws.iter_rows(values_only=True)] for ws in wb.worksheets}
    wb.close()
    return out


def fingerprint():
    db = database.SessionLocal()
    fp = {}
    try:
        for org in db.query(Organization).order_by(Organization.id).all():
            director = (db.query(User)
                        .filter(User.org_id == org.id, User.is_active.is_(True), User.role.in_(DIR_ROLES))
                        .order_by(User.id).first())
            if not director:
                continue
            tok = auth.create_access_token(director.id, org.id, director.role)
            p = f"org{org.id}"

            def get(key, path, **params):
                fp[f"{p} GET {key}"] = canon(_req("GET", path, tok, params or None))

            def xlsx(key, path, **params):
                fp[f"{p} XLSX {key}"] = canon(xlsx_cells(_req("GET", path, tok, params or None, raw=True)))

            # Списки — полностью и детерминированно (id-сорт), без усечения LIMIT.
            fp[f"{p} GET expenses"] = canon(full_expenses(tok))
            get("income", "/api/income", limit=1000)
            get("transfers", "/api/transfers", limit=1000)
            get("users", "/api/users")
            get("dashboard", "/api/dashboard")
            get("departments", "/api/departments")
            get("income_sources", "/api/income-sources")
            get("balances", "/api/reports/balances")
            get("r_summary", "/api/reports/summary", **{"from": FROM, "to": TO})
            get("r_by_employee", "/api/reports/by-employee", **{"from": FROM, "to": TO})
            get("r_by_category", "/api/reports/by-category", **{"from": FROM, "to": TO})
            get("r_balance", "/api/reports/balance", **{"from": FROM, "to": TO})
            for (y, mo) in MONTHS:
                get(f"r_cat_{y}_{mo}", "/api/reports/categories", year=y, month=mo)
                get(f"r_emp_{y}_{mo}", "/api/reports/employees", year=y, month=mo)
                get(f"r_dep_{y}_{mo}", "/api/reports/by-department", year=y, month=mo)
                get(f"r_inc_{y}_{mo}", "/api/reports/incomes", year=y, month=mo)
            xlsx("x_expenses", "/api/expenses/export.xlsx")
            xlsx("x_employees", "/api/reports/employees.xlsx", year=2026, month=8)
            fp[f"{p} POST find_duplicates"] = canon(_req("POST", "/api/admin/find-duplicates", tok, body={}))
            get("recent_ops", "/api/admin/recent-operations", limit=5000)  # без усечения
            for u in db.query(User).filter(User.org_id == org.id).order_by(User.id).all():
                get(f"balance_{u.id}", f"/api/users/{u.id}/balance")
                get(f"profile_{u.id}", f"/api/employees/{u.id}/profile", year=2026, month=8)
    finally:
        db.close()
    return fp


if __name__ == "__main__":
    out_path = sys.argv[1] if len(sys.argv) > 1 else "/root/fp.json"
    data = fingerprint()
    blob = json.dumps(data, sort_keys=True, ensure_ascii=False, default=str)
    digest = hashlib.sha256(blob.encode("utf-8")).hexdigest()
    with open(out_path, "w") as f:
        f.write(blob)
    print(f"keys={len(data)} sha256={digest} -> {out_path}")

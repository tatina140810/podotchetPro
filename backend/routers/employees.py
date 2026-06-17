"""Профиль сотрудника — сводная страница /reports/employees/:id.

Отдельный роутер с префиксом /api/employees. Собирает за период (месяц+год)
пять списков (приходы, передачи, расходы, свои заявки, одобренные им заявки)
плюс summary. Доступ:
  - superadmin/gen_director/auditor/admin → любой сотрудник;
  - accountable → только свой профиль;
  - конфиденциальный (is_confidential) → только superadmin/gen_director и он сам
    (через hidden_user_ids, который исключает me.id).
"""
import io
from datetime import datetime
from decimal import Decimal
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, Query, status
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from sqlalchemy.orm import Session

from auth import get_current_user, is_director_level, is_director_or_auditor
from database import get_db
from models import (
    BalanceTopUp,
    Department,
    EmployeeDepartment,
    Expense,
    Income,
    MoneyRequest,
    User,
)
from services.balance import compute_current_balance, load_org_rates
from services.permissions import hidden_user_ids
from services.plan_limits import ensure_can_export


router = APIRouter(prefix="/api/employees", tags=["employees"])


def _month_range(year: int, month: int) -> tuple[datetime, datetime]:
    start = datetime(year, month, 1)
    end = datetime(year + 1, 1, 1) if month == 12 else datetime(year, month + 1, 1)
    return start, end


def _conv(kgs_val: float, currency: str, usd_rate) -> float:
    """KGS → display currency. Сейчас поддержан только USD (как в reports.py)."""
    if currency == "USD" and usd_rate and usd_rate > 0:
        return round(kgs_val / float(usd_rate), 2)
    return round(kgs_val, 2)


def _load_employee(db: Session, me: User, user_id: int) -> User:
    """Загружает сотрудника с проверкой прав доступа к профилю."""
    u = db.get(User, user_id)
    if not u or u.org_id != me.org_id:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Сотрудник не найден")
    # accountable видит только себя.
    if not is_director_or_auditor(me) and me.id != u.id:
        raise HTTPException(status.HTTP_403_FORBIDDEN, "Нет доступа")
    # Конфиденциальный сотрудник: hidden_user_ids исключает me.id (себя видит всегда),
    # а для admin/auditor вернёт его id → 404.
    if user_id in hidden_user_ids(db, me):
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Сотрудник не найден")
    return u


def _department_name(db: Session, user_id: int) -> Optional[str]:
    rows = (
        db.query(Department.name)
        .join(EmployeeDepartment, EmployeeDepartment.department_id == Department.id)
        .filter(EmployeeDepartment.employee_id == user_id)
        .order_by(Department.name)
        .all()
    )
    names = [n for (n,) in rows]
    return ", ".join(names) if names else None


def _build_profile(db: Session, u: User, month: int, year: int, currency: str) -> dict:
    org_id = u.org_id
    start, end = _month_range(year, month)
    rates = load_org_rates(db, org_id)
    usd_rate = rates.get("USD")

    def kgs(amount, cur) -> float:
        return float(Decimal(str(amount)) * rates.get(cur, Decimal("0")))

    def disp(kgs_val: float) -> float:
        return _conv(kgs_val, currency, usd_rate)

    # --- ПРИХОДЫ: входящие выдачи (topup.user_id=emp) + Income (received_by=emp) ---
    received: list[dict] = []
    for t in (
        db.query(BalanceTopUp)
        .filter(BalanceTopUp.org_id == org_id, BalanceTopUp.user_id == u.id,
                BalanceTopUp.date >= start, BalanceTopUp.date < end)
        .order_by(BalanceTopUp.date.desc())
        .all()
    ):
        received.append({
            "id": t.id, "kind": "topup",
            "date": t.date.isoformat(),
            "from_name": t.admin.name if t.admin else None,
            "from_id": t.admin_id,
            "department_id": t.department_id,
            "amount": float(t.amount), "currency": t.currency,
            "amount_kgs": disp(kgs(t.amount, t.currency)),
            "comment": t.note,
        })
    for i in (
        db.query(Income)
        .filter(Income.org_id == org_id, Income.received_by_id == u.id,
                Income.date >= start, Income.date < end)
        .order_by(Income.date.desc())
        .all()
    ):
        received.append({
            "id": i.id, "kind": "income",
            "date": i.date.isoformat(),
            "from_name": i.source,
            "from_id": None,
            "amount": float(i.amount), "currency": i.currency,
            "amount_kgs": disp(kgs(i.amount, i.currency)),
            "comment": i.description,
        })
    received.sort(key=lambda r: r["date"], reverse=True)

    # --- ПЕРЕДАЛ ДАЛЬШЕ: выдачи где сотрудник — отправитель (admin_id=emp) ---
    transferred: list[dict] = []
    for t in (
        db.query(BalanceTopUp)
        .filter(BalanceTopUp.org_id == org_id, BalanceTopUp.admin_id == u.id,
                BalanceTopUp.date >= start, BalanceTopUp.date < end)
        .order_by(BalanceTopUp.date.desc())
        .all()
    ):
        transferred.append({
            "id": t.id,
            "date": t.date.isoformat(),
            "to_name": t.user.name if t.user else None,
            "to_user_id": t.user_id,
            "department_id": t.department_id,
            "category": t.category.name if t.category else None,
            "category_id": t.category_id,
            "amount": float(t.amount), "currency": t.currency,
            "amount_kgs": disp(kgs(t.amount, t.currency)),
            "comment": t.note,
        })

    # --- РАСХОДЫ: конечные расходы сотрудника (expense_type='expense') ---
    expenses: list[dict] = []
    for e in (
        db.query(Expense)
        .filter(Expense.org_id == org_id, Expense.employee_id == u.id,
                Expense.expense_type == "expense",
                Expense.status.in_(("approved", "pending")),
                Expense.spent_at >= start, Expense.spent_at < end)
        .order_by(Expense.spent_at.desc())
        .all()
    ):
        expenses.append({
            "id": e.id,
            "date": e.spent_at.isoformat(),
            "category": e.category.name if e.category else None,
            "category_id": e.category_id,
            "department_id": e.department_id,
            "amount": float(e.amount), "currency": e.currency,
            "amount_kgs": disp(kgs(e.amount, e.currency)),
            "comment": e.description,
        })

    # --- ЗАЯВКИ: мои (requester) и одобренные мной (approver) ---
    def _req_date(r: MoneyRequest) -> datetime:
        return r.approved_at or r.created_at

    def _req_in_period(r: MoneyRequest) -> bool:
        d = _req_date(r)
        return d is not None and start <= d < end

    requests_own: list[dict] = []
    for r in (
        db.query(MoneyRequest)
        .filter(MoneyRequest.org_id == org_id, MoneyRequest.requester_id == u.id)
        .all()
    ):
        if not _req_in_period(r):
            continue
        requests_own.append({
            "id": r.id,
            "date": _req_date(r).isoformat(),
            "category": r.expense_category.name if r.expense_category else None,
            "amount": float(r.total_amount), "currency": r.currency,
            "amount_kgs": disp(kgs(r.total_amount, r.currency)),
            "status": r.status,
            "comment": r.title,
        })
    requests_own.sort(key=lambda r: r["date"], reverse=True)

    requests_approved_by: list[dict] = []
    for r in (
        db.query(MoneyRequest)
        .filter(MoneyRequest.org_id == org_id, MoneyRequest.approver_id == u.id)
        .all()
    ):
        if not _req_in_period(r):
            continue
        requests_approved_by.append({
            "id": r.id,
            "date": _req_date(r).isoformat(),
            "employee_name": r.requester.name if r.requester else None,
            "category": r.expense_category.name if r.expense_category else None,
            "amount": float(r.total_amount), "currency": r.currency,
            "amount_kgs": disp(kgs(r.total_amount, r.currency)),
            "status": r.status,
            "comment": r.title,
        })
    requests_approved_by.sort(key=lambda r: r["date"], reverse=True)

    # --- SUMMARY (в валюте отображения) ---
    received_total = round(sum(x["amount_kgs"] for x in received), 2)
    transferred_total = round(sum(x["amount_kgs"] for x in transferred), 2)
    spent_total = round(sum(x["amount_kgs"] for x in expenses), 2)
    balance_kgs = float(compute_current_balance(db, org_id, u.id, rates=rates, end=end))
    balance = disp(balance_kgs)
    debt = round(-balance, 2) if balance < 0 else 0.0

    return {
        "employee": {
            "id": u.id,
            "name": u.name,
            "role": u.role,
            "department": _department_name(db, u.id),
            "department_ids": [d.id for d in u.departments],
        },
        "period": {"month": month, "year": year},
        "currency": currency if (currency == "KGS" or usd_rate) else "KGS",
        "summary": {
            "received": {"total": received_total, "count": len(received)},
            "transferred": {"total": transferred_total, "count": len(transferred)},
            "spent": {"total": spent_total, "count": len(expenses)},
            "balance": balance,
            "debt": debt,
        },
        "received": received,
        "transferred": transferred,
        "expenses": expenses,
        "requests_own": requests_own,
        "requests_approved_by": requests_approved_by,
    }


@router.get("/{user_id}/profile")
def employee_profile(
    user_id: int,
    month: int = Query(..., ge=1, le=12),
    year: int = Query(..., ge=2020, le=2100),
    currency: str = Query(default="KGS", pattern="^(KGS|USD)$"),
    db: Session = Depends(get_db),
    me: User = Depends(get_current_user),
):
    u = _load_employee(db, me, user_id)
    data = _build_profile(db, u, month, year, currency)
    # accountable не видит подраздел «Одобрял/отклонял» чужие заявки.
    if not is_director_or_auditor(me):
        data["requests_approved_by"] = []
    return data


@router.get("/{user_id}/profile/export")
def employee_profile_export(
    user_id: int,
    month: int = Query(..., ge=1, le=12),
    year: int = Query(..., ge=2020, le=2100),
    currency: str = Query(default="KGS", pattern="^(KGS|USD)$"),
    db: Session = Depends(get_db),
    me: User = Depends(get_current_user),
):
    # Экспорт — только superadmin/admin/gen_director (director-level).
    if not is_director_level(me):
        raise HTTPException(status.HTTP_403_FORBIDDEN, "Экспорт доступен только директору/админу")
    ensure_can_export(db, me)  # guard плана: экспорт только если can_export
    u = _load_employee(db, me, user_id)
    data = _build_profile(db, u, month, year, currency)
    sym = "$" if data["currency"] == "USD" else "с"

    wb = Workbook()
    head_font = Font(bold=True, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor="4F46E5")

    def _sheet(title: str, headers: list[str], rows: list[list]):
        ws = wb.create_sheet(title[:31])
        ws.append(headers)
        for c in ws[1]:
            c.font = head_font
            c.fill = head_fill
        for r in rows:
            ws.append(r)

    def _d(iso: str) -> str:
        return iso[:10] if iso else ""

    wb.remove(wb.active)  # убираем дефолтный пустой лист

    _sheet("Приходы", ["Дата", "От кого", "Сумма", "Валюта", f"В {sym}", "Комментарий"],
           [[_d(x["date"]), x["from_name"] or "", x["amount"], x["currency"], x["amount_kgs"], x["comment"] or ""]
            for x in data["received"]])
    _sheet("Передал дальше", ["Дата", "Кому", "Категория", "Сумма", "Валюта", f"В {sym}", "Комментарий"],
           [[_d(x["date"]), x["to_name"] or "", x["category"] or "", x["amount"], x["currency"], x["amount_kgs"], x["comment"] or ""]
            for x in data["transferred"]])
    _sheet("Расходы", ["Дата", "Категория", "Сумма", "Валюта", f"В {sym}", "Комментарий"],
           [[_d(x["date"]), x["category"] or "", x["amount"], x["currency"], x["amount_kgs"], x["comment"] or ""]
            for x in data["expenses"]])
    _sheet("Мои заявки", ["Дата", "Категория", "Сумма", "Валюта", "Статус", "Комментарий"],
           [[_d(x["date"]), x["category"] or "", x["amount"], x["currency"], x["status"], x["comment"] or ""]
            for x in data["requests_own"]])
    if data["requests_approved_by"]:
        _sheet("Одобрял", ["Дата", "Сотрудник", "Категория", "Сумма", "Валюта", "Решение", "Комментарий"],
               [[_d(x["date"]), x["employee_name"] or "", x["category"] or "", x["amount"], x["currency"], x["status"], x["comment"] or ""]
                for x in data["requests_approved_by"]])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"profile_{u.name}_{year}_{month:02d}.xlsx".replace(" ", "_")
    from urllib.parse import quote
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(fname)}"},
    )

import io
import os
import uuid
from datetime import datetime
from typing import List, Optional
from urllib.parse import quote

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from sqlalchemy.orm import Session

from auth import (
    get_current_user,
    is_director_level,
    is_director_or_auditor,
    require_auditor,
    require_director_level,
)
from config import get_settings
from database import get_db
from models import (
    BalanceTopUp,
    Category,
    Department,
    EmployeeDepartment,
    EmployeeSpec,
    Expense,
    ExpenseReceipt,
    Income,
    SupplierAdvance,
    SupplierAdvanceTransaction,
    User,
)
from schemas import (
    ExpensePersonalContributionToggle,
    ExpenseCreate,
    ExpenseOut,
    ExpenseReceiptCreate,
    ExpenseReceiptOut,
    ExpenseReview,
    ExpenseUpdate,
)
from services.exchange import get_current_rate
from services.permissions import (
    hidden_user_ids,
    member_active_workspace_id,
    owner_isolation_ws_id,
    visible_user_ids,
    visible_workspace_ids,
    workspace_department_ids,
    workspace_expense_clause,
)
from services.plan_limits import ensure_can_export
from routers.supplier_advances import record_purchase


router = APIRouter(prefix="/api/expenses", tags=["expenses"])
settings = get_settings()

ALLOWED_CONTENT_TYPES = {
    "image/jpeg", "image/png", "image/webp", "image/heic", "image/heif",
    "application/pdf",
    # Офисные документы (Excel/Word) — накладные, акты, реестры.
    "application/vnd.ms-excel",  # .xls
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",  # .xlsx
    "application/msword",  # .doc
    "application/vnd.openxmlformats-officedocument.wordprocessingml.document",  # .docx
    "text/csv",  # .csv
}
# Некоторые браузеры/ОС отдают xls/docx как application/octet-stream — тогда проверяем
# по расширению файла из этого белого списка.
ALLOWED_EXTS = {
    ".jpg", ".jpeg", ".png", ".webp", ".heic", ".heif", ".pdf",
    ".xls", ".xlsx", ".doc", ".docx", ".csv",
}


def _apply_workspace_filter(q, db: Session, me: User):
    """Видимость расходов с учётом проектных пространств.
    - владелец пространства → строго только его пространство (изоляция);
    - остальные → скрыть построчную детализацию чужих пространств."""
    iso = owner_isolation_ws_id(db, me)
    if iso is not None:
        return q.filter(Expense.workspace_id == iso)
    clause = workspace_expense_clause(visible_workspace_ids(db, me))
    if clause is not None:
        q = q.filter(clause)
    return q


def _workspace_blocked(db: Session, me: User, e: Expense) -> bool:
    """True, если расход недоступен me по правилам пространств."""
    iso = owner_isolation_ws_id(db, me)
    if iso is not None:
        return e.workspace_id != iso  # владелец видит только своё пространство
    if e.workspace_id is None:
        return False
    vis = visible_workspace_ids(db, me)
    return vis is not None and e.workspace_id not in vis


def _to_out(e: Expense) -> ExpenseOut:
    out = ExpenseOut.model_validate(e)
    out.employee_name = e.employee.name if e.employee else None
    out.category_name = e.category.name if e.category else None
    out.department_name = e.department.name if e.department else None
    out.recorded_by_name = e.recorded_by.name if e.recorded_by else None
    out.to_user_name = e.to_user.name if e.to_user else None
    out.funded_by_name = e.funded_by.name if e.funded_by else None
    out.supplier_name = e.supplier_advance.supplier_name if e.supplier_advance else None
    return out


@router.get("", response_model=List[ExpenseOut])
def list_expenses(
    employee_id: Optional[int] = None,
    category_id: Optional[int] = None,
    status_filter: Optional[str] = Query(default=None, alias="status", pattern="^(pending|approved|rejected)$"),
    date_from: Optional[datetime] = None,
    date_to: Optional[datetime] = None,
    limit: int = Query(default=200, le=1000),
    db: Session = Depends(get_db),
    me: User = Depends(get_current_user),
):
    q = db.query(Expense).filter(Expense.org_id == me.org_id)
    visible = visible_user_ids(db, me)
    if visible is not None:  # accountable — только свои + подчинённые рекурсивно
        q = q.filter(Expense.employee_id.in_(visible))
    hidden = hidden_user_ids(db, me)
    if hidden:  # Фича 2: расходы конфиденциальных сотрудников полностью скрыты
        q = q.filter(Expense.employee_id.notin_(hidden))
    q = _apply_workspace_filter(q, db, me)  # детализация чужих пространств скрыта
    if employee_id:
        q = q.filter(Expense.employee_id == employee_id)
    if category_id:
        q = q.filter(Expense.category_id == category_id)
    if status_filter:
        q = q.filter(Expense.status == status_filter)
    if date_from:
        q = q.filter(Expense.spent_at >= date_from)
    if date_to:
        q = q.filter(Expense.spent_at < date_to)
    rows = q.order_by(Expense.spent_at.desc()).limit(limit).all()
    return [_to_out(e) for e in rows]


_STATUS_RU = {"pending": "На проверке", "approved": "Принят", "rejected": "Отклонён"}


@router.get("/export.xlsx")
def export_expenses_xlsx(
    employee_id: Optional[int] = None,
    category_id: Optional[int] = None,
    status_filter: Optional[str] = Query(default=None, alias="status", pattern="^(pending|approved|rejected)$"),
    date_from: Optional[datetime] = None,
    date_to: Optional[datetime] = None,
    db: Session = Depends(get_db),
    me: User = Depends(get_current_user),
):
    ensure_can_export(db, me)
    q = db.query(Expense).filter(Expense.org_id == me.org_id)
    visible = visible_user_ids(db, me)
    if visible is not None:
        q = q.filter(Expense.employee_id.in_(visible))
    hidden = hidden_user_ids(db, me)
    if hidden:  # Фича 2: расходы конфиденциальных сотрудников скрыты и в экспорте
        q = q.filter(Expense.employee_id.notin_(hidden))
    q = _apply_workspace_filter(q, db, me)  # детализация чужих пространств скрыта и в экспорте
    if employee_id:
        q = q.filter(Expense.employee_id == employee_id)
    if category_id:
        q = q.filter(Expense.category_id == category_id)
    if status_filter:
        q = q.filter(Expense.status == status_filter)
    if date_from:
        q = q.filter(Expense.spent_at >= date_from)
    if date_to:
        q = q.filter(Expense.spent_at < date_to)
    rows = q.order_by(Expense.spent_at.desc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Расходы"

    headers = ["Дата", "Сотрудник", "Категория", "Сумма", "Валюта", "Описание", "Статус", "Причина отклонения"]
    ws.append(headers)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F46E5")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    for e in rows:
        ws.append([
            e.spent_at.strftime("%d.%m.%Y") if e.spent_at else "",
            e.employee.name if e.employee else "",
            e.category.name if e.category else "",
            float(e.amount) if e.amount is not None else 0,
            e.currency or "",
            e.description or "",
            _STATUS_RU.get(e.status, e.status or ""),
            e.review_comment or "",
        ])

    widths = [12, 22, 20, 14, 8, 40, 14, 30]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    for row_idx in range(2, ws.max_row + 1):
        ws.cell(row=row_idx, column=4).number_format = '#,##0.00'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    fname = f"expenses_{datetime.utcnow().strftime('%Y-%m-%d')}.xlsx"
    headers_resp = {
        "Content-Disposition": f"attachment; filename=\"{fname}\"; filename*=UTF-8''{quote(fname)}",
    }
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers_resp,
    )


@router.post("/upload-receipt")
async def upload_receipt(file: UploadFile = File(...), me: User = Depends(get_current_user)):
    ext = os.path.splitext(file.filename or "")[1].lower()
    if file.content_type not in ALLOWED_CONTENT_TYPES and ext not in ALLOWED_EXTS:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, f"Неподдерживаемый тип: {file.content_type or ext or 'неизвестно'}")

    content = await file.read()
    size_mb = len(content) / 1024 / 1024
    if size_mb > settings.max_upload_mb:
        raise HTTPException(status.HTTP_413_REQUEST_ENTITY_TOO_LARGE, f"Файл больше {settings.max_upload_mb}MB")

    org_dir = os.path.join(settings.upload_dir, str(me.org_id))
    os.makedirs(org_dir, exist_ok=True)
    ext = os.path.splitext(file.filename or "")[1].lower() or ".jpg"
    fname = f"{uuid.uuid4().hex}{ext}"
    path = os.path.join(org_dir, fname)
    with open(path, "wb") as f:
        f.write(content)

    url = f"/uploads/{me.org_id}/{fname}"
    return {"url": url, "size": len(content)}


@router.post("", response_model=ExpenseOut, status_code=status.HTTP_201_CREATED)
def create_expense(
    payload: ExpenseCreate,
    db: Session = Depends(get_db),
    me: User = Depends(get_current_user),
):
    # Режим администратора: вносим за другого пользователя.
    # Только admin. employee_id = on_behalf_of; recorded_by = admin; auto_approve=True
    # (admin сам утверждает в момент ввода — повторного review не нужно).
    on_behalf: Optional[User] = None
    if payload.on_behalf_of_user_id is not None:
        if not is_director_or_auditor(me):
            raise HTTPException(
                status.HTTP_403_FORBIDDEN,
                "Вносить от лица другого может только auditor и выше",
            )
        on_behalf = db.get(User, payload.on_behalf_of_user_id)
        if not on_behalf or on_behalf.org_id != me.org_id:
            raise HTTPException(status.HTTP_400_BAD_REQUEST, "Пользователь on_behalf_of не найден")

    # spec проверяется для того, на чьё имя расход (если on_behalf — берём его spec)
    spec_user_id = on_behalf.id if on_behalf else me.id
    spec = db.query(EmployeeSpec).filter(EmployeeSpec.user_id == spec_user_id).first()

    if spec:
        if spec.allowed_categories and payload.category_id and payload.category_id not in spec.allowed_categories:
            raise HTTPException(status.HTTP_400_BAD_REQUEST, "Эта категория не разрешена")
        if spec.requires_receipt and not payload.receipt_url and not on_behalf:
            # При on_behalf admin может вносить без чека (исторические записи)
            raise HTTPException(status.HTTP_400_BAD_REQUEST, "Требуется фото чека")

    if payload.category_id:
        cat = db.get(Category, payload.category_id)
        if not cat or cat.org_id != me.org_id:
            raise HTTPException(status.HTTP_400_BAD_REQUEST, "Категория не найдена")

    # Подразделение — обязательно. При вводе «от лица» (on_behalf, из профиля сотрудника)
    # можно не указывать: подставим единственное подразделение этого сотрудника.
    department_id = payload.department_id
    if department_id is None and on_behalf is not None:
        dep_ids = [d.id for d in on_behalf.departments]
        if len(dep_ids) == 1:
            department_id = dep_ids[0]
    if department_id is None:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "Укажите подразделение")
    dep = db.get(Department, department_id)
    if not dep or dep.org_id != me.org_id:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "Подразделение не найдено")
    # Подотчётный может выбрать ЛЮБОЕ подразделение своей организации (как и предполагает
    # выпадающий список с all=true и путь редактирования расхода): сотрудник не всегда
    # привязан к подразделению, но расход внести должен. Изоляция проектных пространств
    # обеспечивается отдельно (scoped-список подразделений участника пространства).
    # Проверка «та же org» выше — достаточная граница.

    # Если указан получатель — это «передача» (transfer).
    # Создаём пару: Expense(expense_type='transfer', to_user_id=X) у источника +
    # параллельный BalanceTopUp у получателя на ту же сумму.
    # В compute_current_balance _expenses_approved игнорирует transfer (нет double-count),
    # списание идёт через _topups_out у источника, поступление через _topups_in у получателя.
    # В отчёте «по категориям» transfer-Expense НЕ считаются как расход компании.
    recipient: User | None = None
    if payload.to_user_id is not None:
        if payload.currency != "KGS":
            raise HTTPException(
                status.HTTP_400_BAD_REQUEST,
                "Передавать получателю можно только сомы (KGS)",
            )
        recipient = db.get(User, payload.to_user_id)
        if not recipient or recipient.org_id != me.org_id:
            raise HTTPException(status.HTTP_400_BAD_REQUEST, "Получатель не найден")
        if recipient.id == me.id:
            raise HTTPException(status.HTTP_400_BAD_REQUEST, "Нельзя передать самому себе")
        # Accountable — только своим прямым подотчётным (как в /api/transfers).
        if me.role == "accountable" and recipient.supervisor_id != me.id:
            raise HTTPException(
                status.HTTP_403_FORBIDDEN,
                "Передавать можно только своим подотчётным (вы — их supervisor)",
            )

    # Оплата с депозита у поставщика: источник строго один — баланс сотрудника НЕ трогается
    # (уже списан при внесении аванса). Несовместимо с передачей и «личным вкладом».
    pay_from_advance = payload.payment_source == "supplier_advance"
    if pay_from_advance:
        if payload.supplier_advance_id is None:
            raise HTTPException(status.HTTP_400_BAD_REQUEST, "Укажите депозит поставщика")
        if recipient is not None:
            raise HTTPException(status.HTTP_400_BAD_REQUEST, "Передачу нельзя оплатить с депозита")
        if payload.is_personal_contribution:
            raise HTTPException(status.HTTP_400_BAD_REQUEST, "«Личный вклад» несовместим с оплатой с депозита")

    # Если в описании не указано про получателя — добавим автоматически.
    description = payload.description or ""
    if recipient and "Передано" not in description:
        prefix = f"Передано: {recipient.name}"
        description = f"{prefix}\n{description}".strip() if description else prefix

    # Считаем amount_kgs (КГС-эквивалент). Для KGS = amount. Для USD/RUB = amount × курс.
    # Если курса нет — 400, чтобы расход не "висел в воздухе" без отражения в балансе.
    if payload.currency == "KGS":
        amount_kgs = payload.amount
    else:
        rate = get_current_rate(db, me.org_id, payload.currency, "KGS")
        if rate is None:
            raise HTTPException(
                status.HTTP_400_BAD_REQUEST,
                f"Установите курс {payload.currency}/KGS перед записью расхода в {payload.currency}",
            )
        amount_kgs = payload.amount * rate

    # При on_behalf — auto_approve=True (admin сам утверждает); иначе обычная логика.
    auto_approve = (
        on_behalf is not None
        or is_director_level(me)
        or (spec is not None and not spec.requires_approval)
    )
    employee_id = on_behalf.id if on_behalf else me.id
    # Проектное пространство определяет СЕРВЕР по записи (клиенту не доверяем):
    # если сотрудник — УЧАСТНИК активного пространства (владелец или подотчётный),
    # расход уходит в это пространство.
    workspace_id = member_active_workspace_id(db, employee_id, me.org_id)
    e = Expense(
        org_id=me.org_id,
        employee_id=employee_id,
        advance_id=payload.advance_id,
        category_id=payload.category_id,
        department_id=department_id,
        workspace_id=workspace_id,
        amount=payload.amount,
        currency=payload.currency,
        amount_kgs=amount_kgs,
        description=description or None,
        receipt_url=payload.receipt_url,
        status="approved" if auto_approve else "pending",
        reviewed_by_id=me.id if auto_approve else None,
        recorded_by_id=me.id if on_behalf else None,
        to_user_id=recipient.id if recipient else None,
        expense_type="transfer" if recipient else "expense",
        is_personal_contribution=payload.is_personal_contribution and recipient is None,
        payment_source="supplier_advance" if pay_from_advance else "balance",
        supplier_advance_id=payload.supplier_advance_id if pay_from_advance else None,
        spent_at=payload.spent_at or datetime.utcnow(),
    )
    db.add(e)

    # Списание с депозита поставщика — создаёт purchase-транзакцию (проверяет доступ,
    # валюту и остаток). Баланс сотрудника при этом не меняется (payment_source исключён
    # из _expenses_approved). e.id нужен для привязки транзакции → flush.
    if pay_from_advance:
        db.flush()
        record_purchase(db, me, payload.supplier_advance_id, e)

    # Чек, прикреплённый при создании, дублируем в expense_receipts (единый источник
    # для галереи). receipt_url оставляем для обратной совместимости.
    if payload.receipt_url:
        db.flush()  # нужен e.id
        db.add(ExpenseReceipt(
            org_id=me.org_id,
            expense_id=e.id,
            file_url=payload.receipt_url,
            uploaded_by_id=employee_id,
        ))

    # Парный BalanceTopUp у получателя — для transfer'а.
    if recipient:
        db.add(BalanceTopUp(
            org_id=me.org_id,
            admin_id=employee_id,  # тот, кто оформил передачу (источник денег)
            user_id=recipient.id,
            amount=payload.amount,
            currency=payload.currency,
            amount_kgs=amount_kgs,
            note=f"Из расхода: {payload.description or 'без описания'}",
            date=payload.spent_at or datetime.utcnow(),
            department_id=department_id,
            # Привязка к пространству получателя — деньги попадают в его учёт.
            workspace_id=member_active_workspace_id(db, recipient.id, me.org_id),
        ))

    db.commit()
    db.refresh(e)
    return _to_out(e)


@router.get("/{expense_id}", response_model=ExpenseOut)
def get_expense(expense_id: int, db: Session = Depends(get_db), me: User = Depends(get_current_user)):
    e = db.get(Expense, expense_id)
    if not e or e.org_id != me.org_id:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Не найдено")
    visible = visible_user_ids(db, me)
    if visible is not None and e.employee_id not in visible:
        raise HTTPException(status.HTTP_403_FORBIDDEN, "Нет доступа")
    if e.employee_id in hidden_user_ids(db, me):  # Фича 2: чужой конфиденциальный расход
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Не найдено")
    if _workspace_blocked(db, me, e):  # детализация чужого пространства
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Не найдено")
    return _to_out(e)


@router.patch("/{expense_id}", response_model=ExpenseOut)
def update_expense(
    expense_id: int,
    payload: ExpenseUpdate,
    db: Session = Depends(get_db),
    me: User = Depends(get_current_user),
):
    e = db.get(Expense, expense_id)
    if not e or e.org_id != me.org_id:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Не найдено")
    if not is_director_or_auditor(me) and (e.employee_id != me.id or e.status != "pending"):
        raise HTTPException(status.HTTP_403_FORBIDDEN, "Менять можно только свои pending-расходы")

    data = payload.model_dump(exclude_unset=True)
    if data.get("department_id") is not None:
        dep = db.get(Department, data["department_id"])
        if not dep or dep.org_id != me.org_id:
            raise HTTPException(status.HTTP_400_BAD_REQUEST, "Подразделение не найдено")
    for field, value in data.items():
        setattr(e, field, value)

    # Если изменилась amount или currency — пересчитать KGS-эквивалент по ТЕКУЩЕМУ курсу.
    if "amount" in data or "currency" in data:
        if e.currency == "KGS":
            e.amount_kgs = e.amount
        else:
            from services.exchange import get_current_rate as _rate
            rate = _rate(db, me.org_id, e.currency, "KGS")
            if rate is None:
                raise HTTPException(
                    status.HTTP_400_BAD_REQUEST,
                    f"Курс {e.currency}/KGS не установлен — пересчёт невозможен",
                )
            from decimal import Decimal as _D
            e.amount_kgs = _D(str(e.amount)) * rate

    # Расход оплачен с депозита поставщика и изменилась сумма → синхронизируем
    # purchase-транзакцию и статус депозита (правило: остаток пересчитывается).
    if e.payment_source == "supplier_advance" and e.supplier_advance_id and "amount" in data:
        from decimal import Decimal as _D
        from routers.supplier_advances import advance_remaining, _refresh_status
        tx = (
            db.query(SupplierAdvanceTransaction)
            .filter(
                SupplierAdvanceTransaction.expense_id == e.id,
                SupplierAdvanceTransaction.type == "purchase",
            )
            .first()
        )
        if tx:
            # доступный лимит = текущий остаток + сумма этой же покупки (её «вернём» и спишем заново)
            avail = advance_remaining(db, e.supplier_advance_id) + _D(str(tx.amount))
            if _D(str(e.amount)) > avail:
                raise HTTPException(
                    status.HTTP_400_BAD_REQUEST,
                    f"Новая сумма покупки больше остатка депозита ({avail})",
                )
            tx.amount = e.amount
            db.flush()
            adv = db.get(SupplierAdvance, e.supplier_advance_id)
            if adv:
                _refresh_status(db, adv)

    db.commit()
    db.refresh(e)
    return _to_out(e)


@router.post("/{expense_id}/personal-contribution", response_model=ExpenseOut)
def toggle_personal_contribution(
    expense_id: int,
    payload: ExpensePersonalContributionToggle,
    db: Session = Depends(get_db),
    me: User = Depends(get_current_user),
):
    """Пометить/снять «Расход из личных средств в счёт подразделения».
    Флаг НЕ создаёт приход — он влияет только на агрегат подразделения (такой расход
    считается одновременно приходом и расходом). Личный баланс сотрудника и его
    приходы не трогаются. Позволяет руками проставлять вклады из детализации."""
    e = db.get(Expense, expense_id)
    if not e or e.org_id != me.org_id:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Не найдено")
    if _workspace_blocked(db, me, e):
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Не найдено")
    if not is_director_or_auditor(me) and e.employee_id != me.id:
        raise HTTPException(status.HTTP_403_FORBIDDEN, "Нет доступа")
    if e.expense_type == "transfer":
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "Передача не может быть личным вкладом")
    e.is_personal_contribution = bool(payload.enabled)
    db.commit()
    db.refresh(e)
    return _to_out(e)


@router.post("/{expense_id}/review", response_model=ExpenseOut)
def review_expense(
    expense_id: int,
    payload: ExpenseReview,
    db: Session = Depends(get_db),
    me: User = Depends(require_director_level),
):
    """Бухгалтерская проверка: admin или gen_director выставляет approved/rejected."""
    e = db.get(Expense, expense_id)
    if not e or e.org_id != me.org_id:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Не найдено")
    e.status = payload.status
    e.review_comment = payload.review_comment
    e.reviewed_by_id = me.id
    db.commit()
    db.refresh(e)
    return _to_out(e)


@router.post("/{expense_id}/verify", response_model=ExpenseOut)
def verify_expense(
    expense_id: int,
    db: Session = Depends(get_db),
    me: User = Depends(require_auditor),
):
    """Аудиторская верификация — независимый признак is_verified."""
    e = db.get(Expense, expense_id)
    if not e or e.org_id != me.org_id:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Не найдено")
    e.is_verified = True
    e.verified_by_id = me.id
    db.commit()
    db.refresh(e)
    return _to_out(e)


@router.delete("/{expense_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_expense(expense_id: int, db: Session = Depends(get_db), me: User = Depends(get_current_user)):
    e = db.get(Expense, expense_id)
    if not e or e.org_id != me.org_id:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Не найдено")
    if not is_director_or_auditor(me) and (e.employee_id != me.id or e.status != "pending"):
        raise HTTPException(status.HTTP_403_FORBIDDEN, "Удалять можно только свои pending-расходы")
    # Покупка с депозита: запоминаем депозит, чтобы после удаления (FK CASCADE снесёт
    # purchase-транзакцию) пересчитать статус active/depleted.
    adv_id = e.supplier_advance_id if e.payment_source == "supplier_advance" else None
    db.delete(e)
    db.flush()
    if adv_id:
        from routers.supplier_advances import _refresh_status
        adv = db.get(SupplierAdvance, adv_id)
        if adv:
            _refresh_status(db, adv)
    db.commit()
    return None


# ===================== ЧЕКИ / ДОКУМЕНТЫ РАСХОДА =====================
# Несколько чеков на расход. Файл сначала грузится через POST /upload-receipt
# (возвращает url), затем привязывается сюда. Ключевое правило роли «подотчётный»:
# доклеить чек можно даже ПОСЛЕ проверки (approved/rejected) — менять сам расход нельзя.


def _load_expense_visible(db: Session, me: User, expense_id: int) -> Expense:
    e = db.get(Expense, expense_id)
    if not e or e.org_id != me.org_id:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Не найдено")
    visible = visible_user_ids(db, me)
    if visible is not None and e.employee_id not in visible:
        raise HTTPException(status.HTTP_403_FORBIDDEN, "Нет доступа")
    if e.employee_id in hidden_user_ids(db, me):  # Фича 2
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Не найдено")
    if _workspace_blocked(db, me, e):  # детализация чужого пространства
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Не найдено")
    return e


@router.get("/{expense_id}/receipts", response_model=List[ExpenseReceiptOut])
def list_receipts(expense_id: int, db: Session = Depends(get_db), me: User = Depends(get_current_user)):
    e = _load_expense_visible(db, me, expense_id)
    return (
        db.query(ExpenseReceipt)
        .filter(ExpenseReceipt.expense_id == e.id)
        .order_by(ExpenseReceipt.created_at)
        .all()
    )


@router.post("/{expense_id}/receipts", response_model=ExpenseReceiptOut, status_code=status.HTTP_201_CREATED)
def add_receipt(
    expense_id: int,
    payload: ExpenseReceiptCreate,
    db: Session = Depends(get_db),
    me: User = Depends(get_current_user),
):
    """Прикрепить чек/документ к расходу. Разрешено владельцу или директору
    В ЛЮБОМ статусе расхода — в т.ч. после проверки (дополнительные чеки)."""
    e = db.get(Expense, expense_id)
    if not e or e.org_id != me.org_id:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Не найдено")
    if not is_director_level(me) and e.employee_id != me.id:
        raise HTTPException(status.HTTP_403_FORBIDDEN, "Прикреплять чеки можно только к своим расходам")
    r = ExpenseReceipt(
        org_id=me.org_id,
        expense_id=e.id,
        file_url=payload.file_url,
        file_name=payload.file_name,
        uploaded_by_id=me.id,
    )
    db.add(r)
    db.commit()
    db.refresh(r)
    return r


@router.delete("/{expense_id}/receipts/{receipt_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_receipt(
    expense_id: int,
    receipt_id: int,
    db: Session = Depends(get_db),
    me: User = Depends(get_current_user),
):
    """Удалить чек. Директор — всегда. Владелец — только пока расход pending:
    после проверки чеки только добавляются (чтобы нельзя было подменить)."""
    e = db.get(Expense, expense_id)
    if not e or e.org_id != me.org_id:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Не найдено")
    r = db.get(ExpenseReceipt, receipt_id)
    if not r or r.expense_id != e.id:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Чек не найден")
    allowed = is_director_level(me) or (e.employee_id == me.id and e.status == "pending")
    if not allowed:
        raise HTTPException(
            status.HTTP_403_FORBIDDEN,
            "Удалять чек можно только у своего непроверенного расхода",
        )
    db.delete(r)
    db.commit()
    return None

from datetime import date, datetime, time, timedelta
from decimal import Decimal
from typing import List, Optional

from fastapi import APIRouter, Depends, HTTPException, Query, status
from fastapi.responses import StreamingResponse
from sqlalchemy import func
from sqlalchemy.orm import Session

from auth import (
    get_current_user,
    is_director_or_auditor,
    require_admin,
    require_director_level,
    require_director_or_auditor,
)
from database import get_db
from models import (
    Advance,
    BalanceTopUp,
    Category,
    Department,
    Expense,
    Income,
    Organization,
    User,
)
from schemas import (
    BalanceResponse,
    BalanceRow,
    ByCategoryResponse,
    ByCategoryRow,
    ByEmployeeResponse,
    ByEmployeeRow,
    DayPoint,
    EmployeeSummaryRow,
    ExpenseDetailRow,
    ReportSummary,
    ReportSummaryV2,
)
from services.balance import (
    compute_current_balance,
    issued_total,
    load_org_rates,
    pending_total,
    spent_total,
    to_kgs_expr,
)
from services.excel_export import build_workbook
from services.permissions import hidden_user_ids


router = APIRouter(prefix="/api/reports", tags=["reports"])


# ---------- Параметры периода ----------

def _resolve_period(
    date_from: Optional[date],
    date_to: Optional[date],
) -> tuple[Optional[datetime], Optional[datetime]]:
    """date_from/date_to (YYYY-MM-DD) → datetime начала и конца (exclusive)."""
    start = datetime.combine(date_from, time.min) if date_from else None
    end = datetime.combine(date_to + timedelta(days=1), time.min) if date_to else None
    return start, end


def _resolve_category(db: Session, org_id: int, category: Optional[str]) -> Optional[int]:
    """category может быть int (id) или str (name). Возвращаем int id или None."""
    if not category:
        return None
    try:
        cid = int(category)
        if db.query(Category).filter(Category.id == cid, Category.org_id == org_id).first():
            return cid
        return None
    except (ValueError, TypeError):
        pass
    row = db.query(Category).filter(
        Category.org_id == org_id,
        func.lower(Category.name) == category.strip().lower(),
    ).first()
    return row.id if row else None


# ---------- Старая сводка (для обратной совместимости со старым фронтом) ----------

def _summary_rows(db: Session, org_id: int, employee_ids: Optional[list[int]],
                  date_from: Optional[datetime], date_to: Optional[datetime],
                  currency: Optional[str] = None) -> list[dict]:
    q = db.query(User).filter(User.org_id == org_id)
    if employee_ids:
        q = q.filter(User.id.in_(employee_ids))
    rows = []
    for u in q.all():
        iq = db.query(func.coalesce(func.sum(Advance.amount), 0)).filter(
            Advance.org_id == org_id, Advance.employee_id == u.id,
        )
        if currency:
            iq = iq.filter(Advance.currency == currency)
        if date_from:
            iq = iq.filter(Advance.issued_at >= date_from)
        if date_to:
            iq = iq.filter(Advance.issued_at < date_to)
        issued = Decimal(str(iq.scalar() or 0))

        sq = db.query(func.coalesce(func.sum(Expense.amount), 0)).filter(
            Expense.org_id == org_id, Expense.employee_id == u.id,
            Expense.status.in_(("approved", "pending")),
        )
        if currency:
            sq = sq.filter(Expense.currency == currency)
        if date_from:
            sq = sq.filter(Expense.spent_at >= date_from)
        if date_to:
            sq = sq.filter(Expense.spent_at < date_to)
        spent = Decimal(str(sq.scalar() or 0))

        if issued == 0 and spent == 0:
            continue
        rows.append({
            "employee_id": u.id,
            "employee_name": u.name,
            "issued": issued,
            "spent": spent,
            "balance": issued - spent,
        })
    rows.sort(key=lambda r: r["employee_name"])
    return rows


def _detail_rows(db: Session, org_id: int,
                 employee_ids: Optional[list[int]],
                 category_id: Optional[int],
                 date_from: Optional[datetime],
                 date_to: Optional[datetime],
                 currency: Optional[str] = None) -> list[dict]:
    rows = []

    aq = db.query(Advance).filter(Advance.org_id == org_id)
    if currency:
        aq = aq.filter(Advance.currency == currency)
    if employee_ids:
        aq = aq.filter(Advance.employee_id.in_(employee_ids))
    if date_from:
        aq = aq.filter(Advance.issued_at >= date_from)
    if date_to:
        aq = aq.filter(Advance.issued_at < date_to)
    for a in aq.all():
        rows.append({
            "date": a.issued_at,
            "employee": a.employee.name if a.employee else "—",
            "type": "Выдача",
            "category": "—",
            "amount": a.amount,
            "description": a.purpose or "",
            "status": a.payment_type,
        })

    eq = db.query(Expense).filter(Expense.org_id == org_id)
    if currency:
        eq = eq.filter(Expense.currency == currency)
    if employee_ids:
        eq = eq.filter(Expense.employee_id.in_(employee_ids))
    if category_id:
        eq = eq.filter(Expense.category_id == category_id)
    if date_from:
        eq = eq.filter(Expense.spent_at >= date_from)
    if date_to:
        eq = eq.filter(Expense.spent_at < date_to)
    for e in eq.all():
        rows.append({
            "date": e.spent_at,
            "employee": e.employee.name if e.employee else "—",
            "type": "Расход",
            "category": e.category.name if e.category else "—",
            "amount": e.amount,
            "description": e.description or "",
            "status": e.status,
        })

    rows.sort(key=lambda r: r["date"], reverse=True)
    return rows


def _period_label(date_from: Optional[datetime], date_to: Optional[datetime]) -> str:
    if not date_from and not date_to:
        return "за всё время"
    f = date_from.strftime("%Y-%m-%d") if date_from else "..."
    t = date_to.strftime("%Y-%m-%d") if date_to else "..."
    return f"{f} — {t}"


# ============================================================================
# СЕКЦИЯ 1 — ОБЩИЕ ИТОГИ
# ============================================================================

_ALLOWED_CURRENCIES = ("KGS", "USD", "EUR", "RUB")


def _resolve_currency(currency: Optional[str]) -> str:
    """Валюта отчёта. По умолчанию KGS."""
    if currency and currency.upper() in _ALLOWED_CURRENCIES:
        return currency.upper()
    return "KGS"


@router.get("/summary", response_model=ReportSummaryV2)
def report_summary_v2(
    date_from: Optional[date] = Query(default=None, alias="from"),
    date_to: Optional[date] = Query(default=None, alias="to"),
    employee_id: Optional[int] = None,
    category: Optional[str] = None,
    currency: Optional[str] = None,
    db: Session = Depends(get_db),
    admin: User = Depends(require_admin),
):
    org_id = admin.org_id
    start, end = _resolve_period(date_from, date_to)
    cat_id = _resolve_category(db, org_id, category)
    cur = _resolve_currency(currency)

    # --- Сумма выдач ---
    aq = db.query(func.coalesce(func.sum(Advance.amount), 0)).filter(
        Advance.org_id == org_id, Advance.currency == cur,
    )
    if employee_id is not None:
        aq = aq.filter(Advance.employee_id == employee_id)
    if start:
        aq = aq.filter(Advance.issued_at >= start)
    if end:
        aq = aq.filter(Advance.issued_at < end)
    issued = Decimal(str(aq.scalar() or 0))

    # --- Сумма approved расходов ---
    eq_approved = db.query(func.coalesce(func.sum(Expense.amount), 0)).filter(
        Expense.org_id == org_id,
        Expense.status == "approved",
        Expense.currency == cur,
    )
    if employee_id is not None:
        eq_approved = eq_approved.filter(Expense.employee_id == employee_id)
    if cat_id is not None:
        eq_approved = eq_approved.filter(Expense.category_id == cat_id)
    if start:
        eq_approved = eq_approved.filter(Expense.spent_at >= start)
    if end:
        eq_approved = eq_approved.filter(Expense.spent_at < end)
    spent = Decimal(str(eq_approved.scalar() or 0))

    # --- Pending ---
    eq_pending = db.query(func.coalesce(func.sum(Expense.amount), 0)).filter(
        Expense.org_id == org_id,
        Expense.status == "pending",
        Expense.currency == cur,
    )
    if employee_id is not None:
        eq_pending = eq_pending.filter(Expense.employee_id == employee_id)
    if cat_id is not None:
        eq_pending = eq_pending.filter(Expense.category_id == cat_id)
    if start:
        eq_pending = eq_pending.filter(Expense.spent_at >= start)
    if end:
        eq_pending = eq_pending.filter(Expense.spent_at < end)
    pending = Decimal(str(eq_pending.scalar() or 0))

    # --- Разбивка по дням ---
    by_day: dict[str, dict] = {}

    aq_day = db.query(
        func.date(Advance.issued_at).label("d"),
        func.coalesce(func.sum(Advance.amount), 0).label("s"),
    ).filter(Advance.org_id == org_id, Advance.currency == cur)
    if employee_id is not None:
        aq_day = aq_day.filter(Advance.employee_id == employee_id)
    if start:
        aq_day = aq_day.filter(Advance.issued_at >= start)
    if end:
        aq_day = aq_day.filter(Advance.issued_at < end)
    for d, s in aq_day.group_by("d").all():
        key = str(d)
        by_day.setdefault(key, {"date": key, "issued": Decimal(0), "spent": Decimal(0)})
        by_day[key]["issued"] = Decimal(str(s or 0))

    eq_day = db.query(
        func.date(Expense.spent_at).label("d"),
        func.coalesce(func.sum(Expense.amount), 0).label("s"),
    ).filter(Expense.org_id == org_id, Expense.status == "approved", Expense.currency == cur)
    if employee_id is not None:
        eq_day = eq_day.filter(Expense.employee_id == employee_id)
    if cat_id is not None:
        eq_day = eq_day.filter(Expense.category_id == cat_id)
    if start:
        eq_day = eq_day.filter(Expense.spent_at >= start)
    if end:
        eq_day = eq_day.filter(Expense.spent_at < end)
    for d, s in eq_day.group_by("d").all():
        key = str(d)
        by_day.setdefault(key, {"date": key, "issued": Decimal(0), "spent": Decimal(0)})
        by_day[key]["spent"] = Decimal(str(s or 0))

    day_points = [DayPoint(**v) for v in sorted(by_day.values(), key=lambda x: x["date"])]

    return ReportSummaryV2(
        issued_total=issued,
        spent_total=spent,
        balance=issued - spent,
        pending_total=pending,
        by_day=day_points,
        currency=cur,
    )


# ============================================================================
# СЕКЦИЯ 2 — ПО СОТРУДНИКАМ
# ============================================================================

@router.get("/by-employee", response_model=ByEmployeeResponse)
def report_by_employee(
    date_from: Optional[date] = Query(default=None, alias="from"),
    date_to: Optional[date] = Query(default=None, alias="to"),
    employee_id: Optional[int] = None,
    category: Optional[str] = None,
    currency: Optional[str] = None,
    db: Session = Depends(get_db),
    admin: User = Depends(require_admin),
):
    org_id = admin.org_id
    start, end = _resolve_period(date_from, date_to)
    cat_id = _resolve_category(db, org_id, category)
    cur = _resolve_currency(currency)

    users_q = db.query(User).filter(User.org_id == org_id)
    if employee_id is not None:
        users_q = users_q.filter(User.id == employee_id)

    rows: list[ByEmployeeRow] = []
    for u in users_q.all():
        iq = db.query(func.coalesce(func.sum(Advance.amount), 0)).filter(
            Advance.org_id == org_id,
            Advance.employee_id == u.id,
            Advance.currency == cur,
        )
        if start:
            iq = iq.filter(Advance.issued_at >= start)
        if end:
            iq = iq.filter(Advance.issued_at < end)
        issued = Decimal(str(iq.scalar() or 0))

        approved_q = db.query(func.coalesce(func.sum(Expense.amount), 0)).filter(
            Expense.org_id == org_id,
            Expense.employee_id == u.id,
            Expense.status == "approved",
            Expense.currency == cur,
        )
        if cat_id is not None:
            approved_q = approved_q.filter(Expense.category_id == cat_id)
        if start:
            approved_q = approved_q.filter(Expense.spent_at >= start)
        if end:
            approved_q = approved_q.filter(Expense.spent_at < end)
        spent = Decimal(str(approved_q.scalar() or 0))

        pending_q = db.query(func.coalesce(func.sum(Expense.amount), 0)).filter(
            Expense.org_id == org_id,
            Expense.employee_id == u.id,
            Expense.status == "pending",
            Expense.currency == cur,
        )
        if cat_id is not None:
            pending_q = pending_q.filter(Expense.category_id == cat_id)
        if start:
            pending_q = pending_q.filter(Expense.spent_at >= start)
        if end:
            pending_q = pending_q.filter(Expense.spent_at < end)
        pending = Decimal(str(pending_q.scalar() or 0))

        if issued == 0 and spent == 0 and pending == 0 and employee_id is None:
            continue

        rows.append(ByEmployeeRow(
            employee_id=u.id,
            employee_name=u.name,
            issued=issued,
            spent=spent,
            balance=issued - spent,
            pending=pending,
        ))

    rows.sort(key=lambda r: r.employee_name)

    details = None
    if employee_id is not None:
        dq = db.query(Expense).filter(
            Expense.org_id == org_id,
            Expense.employee_id == employee_id,
            Expense.currency == cur,
        )
        if cat_id is not None:
            dq = dq.filter(Expense.category_id == cat_id)
        if start:
            dq = dq.filter(Expense.spent_at >= start)
        if end:
            dq = dq.filter(Expense.spent_at < end)
        dq = dq.order_by(Expense.spent_at.desc())
        details = [
            ExpenseDetailRow(
                id=e.id,
                spent_at=e.spent_at,
                category_name=e.category.name if e.category else None,
                amount=e.amount,
                description=e.description,
                status=e.status,
                receipt_url=e.receipt_url,
            )
            for e in dq.all()
        ]

    return ByEmployeeResponse(rows=rows, details=details)


# ============================================================================
# СЕКЦИЯ 3 — ПО КАТЕГОРИЯМ
# ============================================================================

@router.get("/by-category", response_model=ByCategoryResponse)
def report_by_category(
    date_from: Optional[date] = Query(default=None, alias="from"),
    date_to: Optional[date] = Query(default=None, alias="to"),
    employee_id: Optional[int] = None,
    category: Optional[str] = None,
    currency: Optional[str] = None,
    db: Session = Depends(get_db),
    admin: User = Depends(require_admin),
):
    org_id = admin.org_id
    start, end = _resolve_period(date_from, date_to)
    cat_id = _resolve_category(db, org_id, category)
    cur = _resolve_currency(currency)

    q = db.query(
        Expense.category_id,
        func.count(Expense.id).label("cnt"),
        func.coalesce(func.sum(Expense.amount), 0).label("sm"),
    ).filter(
        Expense.org_id == org_id,
        Expense.status == "approved",
        Expense.currency == cur,
    )
    if employee_id is not None:
        q = q.filter(Expense.employee_id == employee_id)
    if cat_id is not None:
        q = q.filter(Expense.category_id == cat_id)
    if start:
        q = q.filter(Expense.spent_at >= start)
    if end:
        q = q.filter(Expense.spent_at < end)
    q = q.group_by(Expense.category_id)

    raw = q.all()
    total = sum((Decimal(str(s or 0)) for _, _, s in raw), Decimal(0))

    # имена категорий
    cat_map: dict[int, str] = {
        c.id: c.name
        for c in db.query(Category).filter(Category.org_id == org_id).all()
    }

    rows: list[ByCategoryRow] = []
    for cid, cnt, sm in raw:
        amount = Decimal(str(sm or 0))
        percent = float(amount / total * 100) if total > 0 else 0.0
        rows.append(ByCategoryRow(
            category_id=cid,
            category_name=cat_map.get(cid, "Без категории") if cid else "Без категории",
            operations=int(cnt or 0),
            amount=amount,
            percent=round(percent, 2),
        ))
    rows.sort(key=lambda r: r.amount, reverse=True)

    return ByCategoryResponse(rows=rows, total_amount=total)


# ============================================================================
# СЕКЦИЯ 4 — ОСТАТКИ (за всё время)
# ============================================================================

@router.get("/balances", response_model=BalanceResponse)
def report_balances(
    employee_id: Optional[int] = None,
    currency: Optional[str] = None,
    db: Session = Depends(get_db),
    admin: User = Depends(require_admin),
):
    org_id = admin.org_id
    cur = _resolve_currency(currency)

    users_q = db.query(User).filter(User.org_id == org_id)
    if employee_id is not None:
        users_q = users_q.filter(User.id == employee_id)

    rows: list[BalanceRow] = []
    for u in users_q.all():
        iq = db.query(func.coalesce(func.sum(Advance.amount), 0)).filter(
            Advance.org_id == org_id,
            Advance.employee_id == u.id,
            Advance.currency == cur,
        )
        issued = Decimal(str(iq.scalar() or 0))
        approved_q = db.query(func.coalesce(func.sum(Expense.amount), 0)).filter(
            Expense.org_id == org_id,
            Expense.employee_id == u.id,
            Expense.status == "approved",
            Expense.currency == cur,
        )
        spent = Decimal(str(approved_q.scalar() or 0))
        monthly_limit = Decimal(str(u.spec.monthly_limit)) if u.spec else Decimal(0)
        if issued == 0 and spent == 0 and employee_id is None:
            continue
        rows.append(BalanceRow(
            employee_id=u.id,
            employee_name=u.name,
            issued_total=issued,
            spent_total=spent,
            balance=issued - spent,
            monthly_limit=monthly_limit,
        ))

    rows.sort(key=lambda r: r.employee_name)
    return BalanceResponse(rows=rows)


# ============================================================================
# СТАРЫЙ ЭНДПОИНТ ДЛЯ ОБРАТНОЙ СОВМЕСТИМОСТИ (вернёт ReportSummary как раньше)
# ============================================================================

@router.get("/summary-legacy", response_model=ReportSummary)
def report_summary_legacy(
    employee_ids: Optional[List[int]] = Query(default=None),
    date_from: Optional[datetime] = None,
    date_to: Optional[datetime] = None,
    db: Session = Depends(get_db),
    admin: User = Depends(require_admin),
):
    raw = _summary_rows(db, admin.org_id, employee_ids, date_from, date_to)
    rows = [EmployeeSummaryRow(**r) for r in raw]
    return ReportSummary(
        rows=rows,
        total_issued=sum((r.issued for r in rows), Decimal(0)),
        total_spent=sum((r.spent for r in rows), Decimal(0)),
        total_balance=sum((r.balance for r in rows), Decimal(0)),
    )


# ============================================================================
# EXPORT (GET для прямой ссылки + POST как в ТЗ)
# ============================================================================

def _build_export(
    db: Session,
    admin: User,
    employee_id: Optional[int],
    category: Optional[str],
    currency: Optional[str],
    date_from_dt: Optional[datetime],
    date_to_dt: Optional[datetime],
) -> bytes:
    org = db.get(Organization, admin.org_id)
    employee_ids = [employee_id] if employee_id is not None else None
    cat_id = _resolve_category(db, admin.org_id, category)
    cur = _resolve_currency(currency)
    summary = _summary_rows(db, admin.org_id, employee_ids, date_from_dt, date_to_dt, currency=cur)
    detail = _detail_rows(db, admin.org_id, employee_ids, cat_id, date_from_dt, date_to_dt, currency=cur)
    period = _period_label(date_from_dt, date_to_dt) + f" · {cur}"
    return build_workbook(
        org_name=org.name if org else "—",
        period_label=period,
        summary_rows=summary,
        detail_rows=detail,
    )


def _xlsx_response(blob: bytes) -> StreamingResponse:
    fname = f"podotchet_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        iter([blob]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@router.get("/excel")
def report_excel(
    date_from: Optional[date] = Query(default=None, alias="from"),
    date_to: Optional[date] = Query(default=None, alias="to"),
    employee_id: Optional[int] = None,
    category: Optional[str] = None,
    currency: Optional[str] = None,
    db: Session = Depends(get_db),
    admin: User = Depends(require_admin),
):
    start, end = _resolve_period(date_from, date_to)
    return _xlsx_response(_build_export(db, admin, employee_id, category, currency, start, end))


@router.post("/export")
def report_export(
    date_from: Optional[date] = Query(default=None, alias="from"),
    date_to: Optional[date] = Query(default=None, alias="to"),
    employee_id: Optional[int] = None,
    category: Optional[str] = None,
    currency: Optional[str] = None,
    db: Session = Depends(get_db),
    admin: User = Depends(require_admin),
):
    """POST-версия экспорта (по ТЗ)."""
    start, end = _resolve_period(date_from, date_to)
    return _xlsx_response(_build_export(db, admin, employee_id, category, currency, start, end))


# ===================== /api/reports/categories =====================
# Месячный отчёт: расходы по категориям, по подотчётным, динамика 6 месяцев.

import io as _io
from urllib.parse import quote as _quote
from openpyxl import Workbook as _Workbook
from openpyxl.styles import Font as _Font, PatternFill as _PatternFill


def _month_range(year: int, month: int) -> tuple[datetime, datetime]:
    start = datetime(year, month, 1)
    if month == 12:
        end = datetime(year + 1, 1, 1)
    else:
        end = datetime(year, month + 1, 1)
    return start, end


def _maybe_convert_usd(kgs_value: float, display_currency: str, usd_rate: Optional[Decimal]) -> float:
    """KGS → USD по курсу, если просят USD и курс задан. Иначе остаётся в KGS."""
    if display_currency == "USD" and usd_rate and usd_rate > 0:
        return round(kgs_value / float(usd_rate), 2)
    return round(kgs_value, 2)


def _build_category_report(
    db: Session, org_id: int, year: int, month: int, display_currency: str = "KGS",
    department_id: Optional[int] = None, hidden_ids: Optional[set] = None,
) -> dict:
    """Отчёт «По категориям» — две секции (operational / other), без имён сотрудников.
    Расход = ТОЛЬКО Expense (TopUp не считается расходом компании, это движение).
    При смене курса USD/KGS — пересчитывается через load_org_rates + to_kgs_expr.
    department_id — если задан, считаем только расходы этого подразделения."""
    rates = load_org_rates(db, org_id)
    usd_rate = rates.get("USD")
    start, end = _month_range(year, month)
    expense_expr = to_kgs_expr(Expense.amount, Expense.currency, rates)
    income_expr = to_kgs_expr(Income.amount, Income.currency, rates)

    # --- Все Expense за период с категориями и флагом is_operational ---
    # ВАЖНО: только expense_type='expense' (конечные расходы компании).
    # Transfer-Expense (передачи между подотчётными) НЕ считаются расходом — деньги
    # ещё внутри org. Они учитываются отдельно в блоке by_employee → transferred_out.
    # Также исключаем Expense с категорией is_system=True («Подотчёт») — это внутренний
    # маркер, не реальный расход компании.
    exp_q = (
        db.query(Expense)
        .outerjoin(Category, Category.id == Expense.category_id)
        .filter(
            Expense.org_id == org_id,
            Expense.status.in_(("approved", "pending")),
            Expense.expense_type == "expense",
            Expense.spent_at >= start,
            Expense.spent_at < end,
            (Category.id.is_(None)) | (Category.is_system.is_(False)),
        )
    )
    if department_id is not None:
        exp_q = exp_q.filter(Expense.department_id == department_id)
    if hidden_ids:  # Фича 2: собственные расходы конфиденциальных не считаем
        exp_q = exp_q.filter(Expense.employee_id.notin_(hidden_ids))
    all_expenses = exp_q.order_by(Expense.spent_at.desc()).all()
    categories_map: dict = {
        c.id: c for c in db.query(Category).filter(Category.org_id == org_id).all()
    }
    dept_names: dict = {
        d.id: d.name for d in db.query(Department).filter(Department.org_id == org_id).all()
    }

    # Группируем по category_id
    groups: dict = {}
    for e in all_expenses:
        rate_for_cur = rates.get(e.currency, Decimal("0"))
        kgs_value = float(Decimal(str(e.amount)) * rate_for_cur)
        item = {
            "id": e.id,
            "amount": float(e.amount),
            "currency": e.currency,
            "amount_kgs": kgs_value,
            "description": e.description,
            "spent_at": e.spent_at.isoformat(),
            "status": e.status,
        }
        groups.setdefault(e.category_id, []).append(item)

    operational: list[dict] = []
    other: list[dict] = []
    total_expenses = 0.0
    for cat_id, items in groups.items():
        total_kgs = sum(i["amount_kgs"] for i in items)
        if total_kgs <= 0:
            continue
        cat = categories_map.get(cat_id) if cat_id else None
        cat_name = cat.name if cat else "Без категории"
        is_op = bool(cat.is_operational) if cat else False
        # Колонка «Подразделение»: для общих категорий (department_id=NULL) — пусто.
        cat_dept = dept_names.get(cat.department_id) if (cat and cat.department_id) else None
        row = {
            "category_id": cat_id,
            "category": cat_name,
            "department": cat_dept,
            "amount": _maybe_convert_usd(total_kgs, display_currency, usd_rate),
            "count": len(items),
            "items": [
                {**i, "amount_kgs": _maybe_convert_usd(i["amount_kgs"], display_currency, usd_rate)}
                for i in items
            ],
        }
        total_expenses += total_kgs
        (operational if is_op else other).append(row)

    total_expenses_display = _maybe_convert_usd(total_expenses, display_currency, usd_rate)
    for r in operational + other:
        r["percent"] = round(r["amount"] / total_expenses_display * 100, 1) if total_expenses_display else 0.0
    operational.sort(key=lambda x: x["amount"], reverse=True)
    other.sort(key=lambda x: x["amount"], reverse=True)
    op_subtotal = sum(r["amount"] for r in operational)
    other_subtotal = sum(r["amount"] for r in other)

    # --- Приходы за период ---
    # Приходы (Income) не привязаны к подразделению — при фильтре по подразделению
    # их не показываем (иначе исказят результат конкретного подразделения).
    if department_id is not None:
        total_income_kgs = 0.0
    else:
        total_income_kgs = float(
            db.query(func.coalesce(func.sum(income_expr), 0))
            .filter(Income.org_id == org_id, Income.date >= start, Income.date < end)
            .scalar() or 0
        )
    total_income_display = _maybe_convert_usd(total_income_kgs, display_currency, usd_rate)
    result = total_income_display - total_expenses_display

    return {
        "year": year,
        "month": month,
        "operational": operational,
        "other": other,
        "operational_subtotal": round(op_subtotal, 2),
        "other_subtotal": round(other_subtotal, 2),
        "total_expenses": round(total_expenses_display, 2),
        "total_income": round(total_income_display, 2),
        "result": round(result, 2),
        "currency": display_currency if (display_currency == "KGS" or usd_rate) else "KGS",
        "rate": float(usd_rate) if usd_rate else None,
    }


# ===================== /api/reports/employees =====================
# Отчёт «По сотрудникам»: Получил | Передал дальше | Потратил | Остаток | Долг.
# Остаток = накопительный к концу периода.

def _build_employees_report(
    db: Session, org_id: int, year: int, month: int, display_currency: str = "KGS",
    department_id: Optional[int] = None, hidden_ids: Optional[set] = None,
) -> dict:
    rates = load_org_rates(db, org_id)
    usd_rate = rates.get("USD")
    start, end = _month_range(year, month)
    topup_expr = to_kgs_expr(BalanceTopUp.amount, BalanceTopUp.currency, rates)
    income_expr = to_kgs_expr(Income.amount, Income.currency, rates)
    expense_expr = to_kgs_expr(Expense.amount, Expense.currency, rates)

    # received = TopUp_in + Income (что пришло сотруднику за период)
    topup_in_q = (
        db.query(BalanceTopUp.user_id, func.coalesce(func.sum(topup_expr), 0))
        .filter(BalanceTopUp.org_id == org_id, BalanceTopUp.date >= start, BalanceTopUp.date < end)
    )
    if department_id is not None:
        topup_in_q = topup_in_q.filter(BalanceTopUp.department_id == department_id)
    topup_in_rows = dict(topup_in_q.group_by(BalanceTopUp.user_id).all())
    # Income не привязан к подразделению — при фильтре исключаем.
    if department_id is not None:
        income_in_rows = {}
    else:
        income_in_rows = dict(
            db.query(Income.received_by_id, func.coalesce(func.sum(income_expr), 0))
            .filter(Income.org_id == org_id, Income.date >= start, Income.date < end)
            .group_by(Income.received_by_id).all()
        )
    # transferred_out = TopUp где он admin_id (выдал из своих). С учётом transfer-Expense
    # это уже включает все «передачи» — у каждого transfer есть пара BalanceTopUp.
    topup_out_q = (
        db.query(BalanceTopUp.admin_id, func.coalesce(func.sum(topup_expr), 0))
        .filter(BalanceTopUp.org_id == org_id, BalanceTopUp.date >= start, BalanceTopUp.date < end)
    )
    if department_id is not None:
        topup_out_q = topup_out_q.filter(BalanceTopUp.department_id == department_id)
    topup_out_rows = dict(topup_out_q.group_by(BalanceTopUp.admin_id).all())
    # spent = только конечные Expense (expense_type='expense'). Transfer-Expense НЕ
    # включаем — они уже учтены через topup_out_rows.
    spent_q = (
        db.query(Expense.employee_id, func.coalesce(func.sum(expense_expr), 0))
        .filter(
            Expense.org_id == org_id,
            Expense.status.in_(("approved", "pending")),
            Expense.expense_type == "expense",
            Expense.spent_at >= start,
            Expense.spent_at < end,
        )
    )
    if department_id is not None:
        spent_q = spent_q.filter(Expense.department_id == department_id)
    spent_rows = dict(spent_q.group_by(Expense.employee_id).all())

    users_q = db.query(User).filter(User.org_id == org_id, User.is_active.is_(True))
    if hidden_ids:  # Фича 2: конфиденциальные не показываются в отчёте по сотрудникам
        users_q = users_q.filter(User.id.notin_(hidden_ids))
    users = users_q.all()
    rows: list[dict] = []
    for u in users:
        received_kgs = float(topup_in_rows.get(u.id, 0)) + float(income_in_rows.get(u.id, 0))
        transferred_kgs = float(topup_out_rows.get(u.id, 0))
        spent_kgs = float(spent_rows.get(u.id, 0))
        balance_end_kgs = float(compute_current_balance(db, org_id, u.id, rates=rates, end=end))
        # Пропускаем юзеров без движений и без накопительного остатка
        if received_kgs == 0 and transferred_kgs == 0 and spent_kgs == 0 and abs(balance_end_kgs) < 0.01:
            continue
        received = _maybe_convert_usd(received_kgs, display_currency, usd_rate)
        transferred = _maybe_convert_usd(transferred_kgs, display_currency, usd_rate)
        spent = _maybe_convert_usd(spent_kgs, display_currency, usd_rate)
        balance = _maybe_convert_usd(balance_end_kgs, display_currency, usd_rate)
        debt = -balance if balance < 0 else 0.0
        rows.append({
            "user_id": u.id,
            "name": u.name,
            "received": received,
            "transferred_out": transferred,
            "spent": spent,
            "balance": balance,
            "debt": round(debt, 2),
        })
    # Сортируем: сначала с наибольшим долгом, потом по received убыванию
    rows.sort(key=lambda r: (-r["debt"], -r["received"]))

    return {
        "year": year,
        "month": month,
        "rows": rows,
        "currency": display_currency if (display_currency == "KGS" or usd_rate) else "KGS",
        "rate": float(usd_rate) if usd_rate else None,
    }


# ===================== /api/reports/balance =====================
# Хронологическая лента всех операций org. Накопительный остаток org = Income − Expense.

def _build_balance_report(
    db: Session, org_id: int, date_from: datetime, date_to: datetime, display_currency: str = "KGS",
    department_id: Optional[int] = None, hidden_ids: Optional[set] = None,
) -> dict:
    rates = load_org_rates(db, org_id)
    usd_rate = rates.get("USD")
    hidden = hidden_ids or set()

    def _kgs(amount, currency):
        return float(Decimal(str(amount)) * rates.get(currency, Decimal("0")))

    # Накопительный остаток ДО date_from (Income − Expense за всю историю до начала периода)
    income_expr = to_kgs_expr(Income.amount, Income.currency, rates)
    expense_expr = to_kgs_expr(Expense.amount, Expense.currency, rates)

    # При фильтре по подразделению приходы (Income, без подразделения) не учитываем.
    if department_id is not None:
        income_before = 0.0
    else:
        income_before_q = (
            db.query(func.coalesce(func.sum(income_expr), 0))
            .filter(Income.org_id == org_id, Income.date < date_from)
        )
        if hidden:  # Фича 2: приходы конфиденциальных не учитываем
            income_before_q = income_before_q.filter(Income.received_by_id.notin_(hidden))
        income_before = float(income_before_q.scalar() or 0)
    exp_before_q = (
        db.query(func.coalesce(func.sum(expense_expr), 0))
        .filter(
            Expense.org_id == org_id,
            Expense.status.in_(("approved", "pending")),
            Expense.spent_at < date_from,
        )
    )
    if department_id is not None:
        exp_before_q = exp_before_q.filter(Expense.department_id == department_id)
    if hidden:  # Фича 2: собственные расходы конфиденциальных не учитываем
        exp_before_q = exp_before_q.filter(Expense.employee_id.notin_(hidden))
    expense_before = float(exp_before_q.scalar() or 0)
    opening_balance_kgs = income_before - expense_before

    # Все операции за период
    operations: list[dict] = []
    exp_period_q = (
        db.query(Expense)
        .filter(
            Expense.org_id == org_id,
            Expense.status.in_(("approved", "pending")),
            Expense.spent_at >= date_from,
            Expense.spent_at < date_to,
        )
    )
    if department_id is not None:
        exp_period_q = exp_period_q.filter(Expense.department_id == department_id)
    if hidden:  # Фича 2
        exp_period_q = exp_period_q.filter(Expense.employee_id.notin_(hidden))
    for e in exp_period_q.all():
        kgs = _kgs(e.amount, e.currency)
        operations.append({
            "type": "expense",
            "id": e.id,
            "date": e.spent_at.date().isoformat(),
            "datetime": e.spent_at.isoformat(),
            "amount": float(e.amount),
            "currency": e.currency,
            "amount_kgs": kgs,
            "description": e.description,
            "category": e.category.name if e.category else "Без категории",
            "who": e.employee.name if e.employee else None,
            "color": "red",
        })
    if department_id is not None:
        income_rows = []
    else:
        income_q = db.query(Income).filter(
            Income.org_id == org_id, Income.date >= date_from, Income.date < date_to
        )
        if hidden:  # Фича 2: приходы конфиденциальных скрыты
            income_q = income_q.filter(Income.received_by_id.notin_(hidden))
        income_rows = income_q.all()
    for i in income_rows:
        kgs = _kgs(i.amount, i.currency)
        operations.append({
            "type": "income",
            "id": i.id,
            "date": i.date.date().isoformat(),
            "datetime": i.date.isoformat(),
            "amount": float(i.amount),
            "currency": i.currency,
            "amount_kgs": kgs,
            "description": i.description or i.source,
            "source": i.source,
            "who": i.received_by.name if i.received_by else None,
            "color": "green",
        })
    topup_period_q = (
        db.query(BalanceTopUp)
        .filter(BalanceTopUp.org_id == org_id, BalanceTopUp.date >= date_from, BalanceTopUp.date < date_to)
    )
    if department_id is not None:
        topup_period_q = topup_period_q.filter(BalanceTopUp.department_id == department_id)
    if hidden:  # Фича 2: выдачи, где конфиденциальный — отправитель ИЛИ получатель, скрыты
        topup_period_q = topup_period_q.filter(
            BalanceTopUp.admin_id.notin_(hidden),
            BalanceTopUp.user_id.notin_(hidden),
        )
    for t in topup_period_q.all():
        kgs = _kgs(t.amount, t.currency)
        operations.append({
            "type": "transfer",
            "id": t.id,
            "date": t.date.date().isoformat(),
            "datetime": t.date.isoformat(),
            "amount": float(t.amount),
            "currency": t.currency,
            "amount_kgs": kgs,
            "description": t.note,
            "category": t.category.name if t.category else None,
            "from": t.admin.name if t.admin else None,
            "to": t.user.name if t.user else None,
            "color": "gray",
        })

    # Сортируем по datetime
    operations.sort(key=lambda o: o["datetime"])

    # Группируем по дням
    by_day: dict[str, dict] = {}
    for op in operations:
        day = op["date"]
        by_day.setdefault(day, {
            "date": day,
            "operations": [],
            "day_income_kgs": 0.0,
            "day_expense_kgs": 0.0,
        })
        by_day[day]["operations"].append(op)
        if op["type"] == "income":
            by_day[day]["day_income_kgs"] += op["amount_kgs"]
        elif op["type"] == "expense":
            by_day[day]["day_expense_kgs"] += op["amount_kgs"]
        # transfer не влияет на org-баланс

    # Накопительный остаток по дням
    days: list[dict] = []
    running_kgs = opening_balance_kgs
    for day_key in sorted(by_day.keys()):
        d = by_day[day_key]
        day_result_kgs = d["day_income_kgs"] - d["day_expense_kgs"]
        running_kgs += day_result_kgs
        # Конвертация для отображения
        d_ops = []
        for op in d["operations"]:
            d_ops.append({**op, "amount_kgs": _maybe_convert_usd(op["amount_kgs"], display_currency, usd_rate)})
        days.append({
            "date": d["date"],
            "operations": d_ops,
            "day_result": _maybe_convert_usd(day_result_kgs, display_currency, usd_rate),
            "cumulative_balance": _maybe_convert_usd(running_kgs, display_currency, usd_rate),
        })

    period_income_kgs = sum(d["day_income_kgs"] for d in by_day.values())
    period_expense_kgs = sum(d["day_expense_kgs"] for d in by_day.values())

    return {
        "from": date_from.date().isoformat(),
        "to": (date_to - timedelta(seconds=1)).date().isoformat(),
        "opening_balance": _maybe_convert_usd(opening_balance_kgs, display_currency, usd_rate),
        "closing_balance": _maybe_convert_usd(running_kgs, display_currency, usd_rate),
        "days": days,
        "period_total": {
            "income": _maybe_convert_usd(period_income_kgs, display_currency, usd_rate),
            "expenses": _maybe_convert_usd(period_expense_kgs, display_currency, usd_rate),
            "result": _maybe_convert_usd(period_income_kgs - period_expense_kgs, display_currency, usd_rate),
        },
        "currency": display_currency if (display_currency == "KGS" or usd_rate) else "KGS",
        "rate": float(usd_rate) if usd_rate else None,
    }


@router.get("/categories")
def category_report(
    year: int = Query(..., ge=2020, le=2100),
    month: int = Query(..., ge=1, le=12),
    currency: str = Query(default="KGS", pattern="^(KGS|USD)$"),
    department_id: Optional[int] = None,
    db: Session = Depends(get_db),
    me: User = Depends(require_director_or_auditor),
):
    return _build_category_report(
        db, me.org_id, year, month, currency,
        department_id=department_id, hidden_ids=hidden_user_ids(db, me),
    )


@router.get("/employees")
def employees_report(
    year: int = Query(..., ge=2020, le=2100),
    month: int = Query(..., ge=1, le=12),
    currency: str = Query(default="KGS", pattern="^(KGS|USD)$"),
    department_id: Optional[int] = None,
    db: Session = Depends(get_db),
    me: User = Depends(require_director_or_auditor),
):
    return _build_employees_report(
        db, me.org_id, year, month, currency,
        department_id=department_id, hidden_ids=hidden_user_ids(db, me),
    )


@router.get("/incomes")
def incomes_report(
    year: Optional[int] = Query(default=None, ge=2020, le=2100),
    month: Optional[int] = Query(default=None, ge=1, le=12),
    currency: str = Query(default="KGS", pattern="^(KGS|USD)$"),
    db: Session = Depends(get_db),
    me: User = Depends(require_director_level),
):
    """Полный отчёт о приходах org за период (admin + gen_director).
    Возвращает список Income с датами, источниками, описанием, получателем,
    суммами в родной валюте + KGS-эквиваленте по текущему курсу."""
    rates = load_org_rates(db, me.org_id)
    usd_rate = rates.get("USD")
    # Период: если year+month заданы — окно месяца; иначе все.
    start: Optional[datetime] = None
    end: Optional[datetime] = None
    if year is not None and month is not None:
        start, end = _month_range(year, month)
    q = db.query(Income).filter(Income.org_id == me.org_id)
    if start:
        q = q.filter(Income.date >= start)
    if end:
        q = q.filter(Income.date < end)
    rows = q.order_by(Income.date.desc()).all()

    items: list[dict] = []
    total_kgs = 0.0
    for i in rows:
        rate_for_cur = rates.get(i.currency, Decimal("0"))
        kgs_value = float(Decimal(str(i.amount)) * rate_for_cur)
        total_kgs += kgs_value
        items.append({
            "id": i.id,
            "date": i.date.isoformat(),
            "amount": float(i.amount),
            "currency": i.currency,
            "amount_kgs": kgs_value,
            "amount_display": _maybe_convert_usd(kgs_value, currency, usd_rate),
            "source": i.source,
            "description": i.description,
            "received_by_name": i.received_by.name if i.received_by else None,
            "received_by_id": i.received_by_id,
            "created_by_name": i.created_by.name if i.created_by else None,
        })
    return {
        "year": year,
        "month": month,
        "items": items,
        "total": _maybe_convert_usd(total_kgs, currency, usd_rate),
        "count": len(items),
        "currency": currency if (currency == "KGS" or usd_rate) else "KGS",
        "rate": float(usd_rate) if usd_rate else None,
    }


@router.get("/balance")
def balance_report(
    date_from: Optional[date] = Query(default=None, alias="from"),
    date_to: Optional[date] = Query(default=None, alias="to"),
    currency: str = Query(default="KGS", pattern="^(KGS|USD)$"),
    department_id: Optional[int] = None,
    db: Session = Depends(get_db),
    me: User = Depends(require_director_or_auditor),
):
    # По умолчанию — текущий месяц
    today = datetime.utcnow()
    if date_from is None:
        date_from = date(today.year, today.month, 1)
    if date_to is None:
        # включительно — переводим в exclusive datetime ниже
        if today.month == 12:
            date_to = date(today.year + 1, 1, 1)
        else:
            date_to = date(today.year, today.month + 1, 1)
        date_to = date_to - timedelta(days=1)
    df = datetime.combine(date_from, time.min)
    dt = datetime.combine(date_to + timedelta(days=1), time.min)
    return _build_balance_report(
        db, me.org_id, df, dt, currency,
        department_id=department_id, hidden_ids=hidden_user_ids(db, me),
    )


# ===================== Excel-экспорт =====================

@router.get("/categories.xlsx")
def category_report_xlsx(
    year: int = Query(..., ge=2020, le=2100),
    month: int = Query(..., ge=1, le=12),
    currency: str = Query(default="KGS", pattern="^(KGS|USD)$"),
    db: Session = Depends(get_db),
    me: User = Depends(require_director_or_auditor),
):
    data = _build_category_report(db, me.org_id, year, month, currency, hidden_ids=hidden_user_ids(db, me))
    sym = "$" if data["currency"] == "USD" else "с"

    wb = _Workbook()
    ws = wb.active
    ws.title = "Отчёт по категориям"

    bold_white = _Font(bold=True, color="FFFFFF")
    op_fill = _PatternFill("solid", fgColor="4F46E5")
    other_fill = _PatternFill("solid", fgColor="6B7280")
    section_fill = _PatternFill("solid", fgColor="E0E7FF")

    def header_row(title: str, fill):
        ws.append([title, "Сумма", "%", "Операций"])
        for cell in ws[ws.max_row]:
            cell.font = bold_white
            cell.fill = fill

    header_row("ОПЕРАЦИОННЫЕ РАСХОДЫ", op_fill)
    for r in data["operational"]:
        ws.append([r["category"], r["amount"], f"{r['percent']}%", r["count"]])
    ws.append(["Итого операционные:", data["operational_subtotal"], "", ""])
    for cell in ws[ws.max_row]:
        cell.font = _Font(bold=True)
        cell.fill = section_fill
    ws.append([])

    header_row("ПРОЧИЕ РАСХОДЫ", other_fill)
    for r in data["other"]:
        ws.append([r["category"], r["amount"], f"{r['percent']}%", r["count"]])
    ws.append(["Итого прочие:", data["other_subtotal"], "", ""])
    for cell in ws[ws.max_row]:
        cell.font = _Font(bold=True)
        cell.fill = section_fill
    ws.append([])

    ws.append([f"ИТОГО РАСХОД ({sym}):", data["total_expenses"], "", ""])
    ws.append([f"ИТОГО ПРИХОД ({sym}):", data["total_income"], "", ""])
    ws.append([f"РЕЗУЛЬТАТ ({sym}):", data["result"], "", ""])
    for row_idx in (ws.max_row - 2, ws.max_row - 1, ws.max_row):
        for cell in ws[row_idx]:
            cell.font = _Font(bold=True, size=12)
    for col_idx, w in enumerate([34, 16, 10, 12], start=1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = w

    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"report_categories_{year}_{month:02d}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename=\"{fname}\"; filename*=UTF-8''{_quote(fname)}",
        },
    )


@router.get("/employees.xlsx")
def employees_report_xlsx(
    year: int = Query(..., ge=2020, le=2100),
    month: int = Query(..., ge=1, le=12),
    currency: str = Query(default="KGS", pattern="^(KGS|USD)$"),
    db: Session = Depends(get_db),
    me: User = Depends(require_director_or_auditor),
):
    data = _build_employees_report(db, me.org_id, year, month, currency, hidden_ids=hidden_user_ids(db, me))
    sym = "$" if data["currency"] == "USD" else "с"
    wb = _Workbook()
    ws = wb.active
    ws.title = "По сотрудникам"
    ws.append(["Сотрудник", "Получил", "Передал дальше", "Потратил", "Остаток", "Долг"])
    for cell in ws[1]:
        cell.font = _Font(bold=True, color="FFFFFF")
        cell.fill = _PatternFill("solid", fgColor="4F46E5")
    for r in data["rows"]:
        ws.append([r["name"], r["received"], r["transferred_out"], r["spent"], r["balance"], r["debt"]])
    ws.append([])
    ws.append([f"Валюта: {sym}", "", "", "", "", ""])
    for col_idx, w in enumerate([24, 14, 18, 14, 14, 14], start=1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = w
    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"report_employees_{year}_{month:02d}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=\"{fname}\"; filename*=UTF-8''{_quote(fname)}"},
    )


# Развёртка по одному сотруднику: все его операции за месяц (как в раскрывающейся строке таблицы)
_KIND_RU = {
    "topup": "Выдача (получил)",
    "topup_out": "Выдача (отдал)",
    "income": "Приход",
    "transfer_in": "Получил перевод",
    "transfer_out": "Передал перевод",
    "request_approved": "Заявка (получено)",
    "request_approved_out": "Заявка (выдал)",
    "expense": "Расход",
}


@router.get("/employees/{user_id}/details.xlsx")
def employee_details_xlsx(
    user_id: int,
    year: int = Query(..., ge=2020, le=2100),
    month: int = Query(..., ge=1, le=12),
    db: Session = Depends(get_db),
    me: User = Depends(require_director_or_auditor),
):
    """Excel-файл с развёрткой операций сотрудника за выбранный месяц.
    Используется кнопкой рядом с каждым сотрудником в /reports/employees."""
    from routers.users import build_user_history_entries  # избежать циклического импорта при загрузке модуля

    u = db.get(User, user_id)
    if not u or u.org_id != me.org_id:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Сотрудник не найден")
    if user_id in hidden_user_ids(db, me):  # Фича 2: выгрузка конфиденциального скрыта
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Сотрудник не найден")

    start, end = _month_range(year, month)
    entries = build_user_history_entries(db, me.org_id, u.id, start, end)

    wb = _Workbook()
    ws = wb.active
    ws.title = (u.name or f"user_{u.id}")[:31]  # лимит названия листа в Excel
    ws.append(["Дата", "Тип", "Кто/Категория", "Сумма", "Валюта", "Описание"])
    for cell in ws[1]:
        cell.font = _Font(bold=True, color="FFFFFF")
        cell.fill = _PatternFill("solid", fgColor="4F46E5")
    for e in entries:
        ws.append([
            e.created_at.strftime("%Y-%m-%d") if e.created_at else "",
            _KIND_RU.get(e.kind, e.kind),
            e.counterparty or "",
            float(e.amount),
            e.currency or "KGS",
            e.note or "",
        ])
    # Итог: сумма по валютам (раздельно, без конвертации — это исходные суммы)
    if entries:
        ws.append([])
        totals: dict[str, Decimal] = {}
        for e in entries:
            cur = e.currency or "KGS"
            totals[cur] = totals.get(cur, Decimal(0)) + Decimal(str(e.amount))
        for cur, total in totals.items():
            ws.append(["", "ИТОГО", "", float(total), cur, ""])
        for cell in ws[ws.max_row - len(totals) + 1: ws.max_row + 1]:
            for c in cell:
                c.font = _Font(bold=True)
    for col_idx, w in enumerate([12, 22, 24, 14, 8, 40], start=1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = w

    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    # filename="..." требует latin-1, поэтому ASCII-fallback по user_id.
    # filename*=UTF-8''... — реальное имя сотрудника (кириллица URL-кодируется).
    ascii_fname = f"employee_{u.id}_{year}_{month:02d}.xlsx"
    pretty_name = (u.name or f"user{u.id}").replace('"', "'")
    utf8_fname = f"{pretty_name}_{year}_{month:02d}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=\"{ascii_fname}\"; filename*=UTF-8''{_quote(utf8_fname)}"},
    )


@router.get("/employees/{user_id}/history.xlsx")
def employee_history_xlsx(
    user_id: int,
    date_from: Optional[datetime] = None,
    date_to: Optional[datetime] = None,
    db: Session = Depends(get_db),
    me: User = Depends(get_current_user),
):
    """Excel со ВСЕЙ историей движения денег сотрудника (приходы + расходы + переводы).
    Доступ: сам сотрудник (для своей истории) ИЛИ директор/аудитор. В отличие от
    details.xlsx — без обязательного месяца (можно весь период или произвольный range),
    и доступен самому подотчётному."""
    from routers.users import build_user_history_entries  # избегаем циклического импорта

    u = db.get(User, user_id)
    if not u or u.org_id != me.org_id:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Сотрудник не найден")
    if me.id != u.id and not is_director_or_auditor(me):
        raise HTTPException(status.HTTP_403_FORBIDDEN, "Нет доступа")
    if user_id in hidden_user_ids(db, me):  # Фича 2 (сам сотрудник видит себя)
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Сотрудник не найден")

    entries = build_user_history_entries(db, me.org_id, u.id, date_from, date_to)

    wb = _Workbook()
    ws = wb.active
    ws.title = (u.name or f"user_{u.id}")[:31]
    ws.append(["Дата", "Тип", "Кто/Категория", "Сумма", "Валюта", "Описание"])
    for cell in ws[1]:
        cell.font = _Font(bold=True, color="FFFFFF")
        cell.fill = _PatternFill("solid", fgColor="4F46E5")
    for e in entries:
        ws.append([
            e.created_at.strftime("%Y-%m-%d") if e.created_at else "",
            _KIND_RU.get(e.kind, e.kind),
            e.counterparty or "",
            float(e.amount),
            e.currency or "KGS",
            e.note or "",
        ])
    if entries:
        ws.append([])
        totals: dict[str, Decimal] = {}
        for e in entries:
            cur = e.currency or "KGS"
            totals[cur] = totals.get(cur, Decimal(0)) + Decimal(str(e.amount))
        for cur, total in totals.items():
            ws.append(["", "ИТОГО", "", float(total), cur, ""])
        for cell in ws[ws.max_row - len(totals) + 1: ws.max_row + 1]:
            for c in cell:
                c.font = _Font(bold=True)
    for col_idx, w in enumerate([12, 22, 24, 14, 8, 40], start=1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = w

    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    ascii_fname = f"history_{u.id}.xlsx"
    pretty_name = (u.name or f"user{u.id}").replace('"', "'")
    utf8_fname = f"История_{pretty_name}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=\"{ascii_fname}\"; filename*=UTF-8''{_quote(utf8_fname)}"},
    )

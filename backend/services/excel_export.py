"""Генератор Excel-отчёта: 2 листа (Сводный + Детальный) с форматированием."""
from datetime import datetime
from decimal import Decimal
from io import BytesIO
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill("solid", fgColor="2D2A4F")
ALT_FILL = PatternFill("solid", fgColor="F4F2FB")
TOTAL_FILL = PatternFill("solid", fgColor="E8E5F7")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TOTAL_FONT = Font(bold=True, size=11)
THIN = Side(border_style="thin", color="D0D0D0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")


def _autosize(ws, min_w: int = 10, max_w: int = 50) -> None:
    for col_idx, col in enumerate(ws.columns, start=1):
        length = min_w
        for cell in col:
            v = "" if cell.value is None else str(cell.value)
            if len(v) > length:
                length = len(v)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(length + 2, max_w)


def _write_header(ws, row: int, headers: Iterable[str]) -> None:
    for col, name in enumerate(headers, start=1):
        c = ws.cell(row=row, column=col, value=name)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = CENTER
        c.border = BORDER


def build_workbook(
    org_name: str,
    period_label: str,
    summary_rows: list[dict],
    detail_rows: list[dict],
) -> bytes:
    wb = Workbook()

    # ===== Лист 1: Сводный =====
    ws1 = wb.active
    ws1.title = "Сводный"

    ws1.cell(row=1, column=1, value=f"PodotchetPRO — Сводный отчёт").font = Font(bold=True, size=14)
    ws1.cell(row=2, column=1, value=f"Организация: {org_name}")
    ws1.cell(row=3, column=1, value=f"Период: {period_label}")
    ws1.cell(row=4, column=1, value=f"Сформирован: {datetime.now().strftime('%Y-%m-%d %H:%M')}")

    headers1 = ["Сотрудник", "Выдано", "Потрачено", "Остаток"]
    _write_header(ws1, row=6, headers=headers1)

    total_issued = Decimal(0)
    total_spent = Decimal(0)
    total_balance = Decimal(0)

    for i, r in enumerate(summary_rows):
        row = 7 + i
        fill = ALT_FILL if i % 2 else None
        cells = [
            ws1.cell(row=row, column=1, value=r["employee_name"]),
            ws1.cell(row=row, column=2, value=float(r["issued"])),
            ws1.cell(row=row, column=3, value=float(r["spent"])),
            ws1.cell(row=row, column=4, value=float(r["balance"])),
        ]
        for c in cells:
            c.border = BORDER
            if fill:
                c.fill = fill
        cells[0].alignment = LEFT
        for c in cells[1:]:
            c.alignment = RIGHT
            c.number_format = "#,##0.00"

        total_issued += Decimal(str(r["issued"]))
        total_spent += Decimal(str(r["spent"]))
        total_balance += Decimal(str(r["balance"]))

    total_row = 7 + len(summary_rows)
    ws1.cell(row=total_row, column=1, value="ИТОГО").font = TOTAL_FONT
    for col, val in enumerate([total_issued, total_spent, total_balance], start=2):
        c = ws1.cell(row=total_row, column=col, value=float(val))
        c.font = TOTAL_FONT
        c.fill = TOTAL_FILL
        c.alignment = RIGHT
        c.number_format = "#,##0.00"
        c.border = BORDER
    ws1.cell(row=total_row, column=1).fill = TOTAL_FILL
    ws1.cell(row=total_row, column=1).border = BORDER

    _autosize(ws1)
    ws1.freeze_panes = "A7"

    # ===== Лист 2: Детальный =====
    ws2 = wb.create_sheet(title="Детальный")
    ws2.cell(row=1, column=1, value="PodotchetPRO — Детальный отчёт").font = Font(bold=True, size=14)
    ws2.cell(row=2, column=1, value=f"Организация: {org_name}")
    ws2.cell(row=3, column=1, value=f"Период: {period_label}")

    headers2 = ["Дата", "Сотрудник", "Тип", "Категория", "Сумма", "Описание", "Статус"]
    _write_header(ws2, row=5, headers=headers2)

    for i, r in enumerate(detail_rows):
        row = 6 + i
        fill = ALT_FILL if i % 2 else None
        cells = [
            ws2.cell(row=row, column=1, value=r["date"].strftime("%Y-%m-%d") if hasattr(r["date"], "strftime") else r["date"]),
            ws2.cell(row=row, column=2, value=r["employee"]),
            ws2.cell(row=row, column=3, value=r["type"]),
            ws2.cell(row=row, column=4, value=r.get("category") or "—"),
            ws2.cell(row=row, column=5, value=float(r["amount"])),
            ws2.cell(row=row, column=6, value=r.get("description") or ""),
            ws2.cell(row=row, column=7, value=r.get("status") or "—"),
        ]
        for c in cells:
            c.border = BORDER
            if fill:
                c.fill = fill
        cells[4].alignment = RIGHT
        cells[4].number_format = "#,##0.00"

    _autosize(ws2)
    ws2.freeze_panes = "A6"

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()

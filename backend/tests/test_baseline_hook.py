"""Baseline-проверка глобального soft-delete хука (шаг 2 плана, регресс после шага 3).

1) Нулевой дифф под РАЗНЫМИ ролями (director/accountable/auditor/ws-member/
   confidential) — хук ничего не меняет на данных без удалённых. Разные роли = разные
   ветки запросов (permissions.py, workspaces.py), поэтому проверяем каждую.
2) Хук реально применяется на РАЗНЫХ типах записей (личный вклад, депозит, передача,
   USD, запись в пространстве) — soft-delete убирает их из агрегатов.
Схема поднимается миграциями (conftest), не create_all.
"""
import io
import json
import time
from decimal import Decimal

import pytest
from openpyxl import load_workbook

import database
from conftest import auth_headers, disable_hook, enable_hook, reset_all
from seed import seed_all
from services.balance import (
    compute_balances_by_currency,
    compute_current_balance,
    compute_total_issued,
    compute_total_received,
    pending_total,
    spent_total,
)

FROM, TO = "2026-08-01", "2026-08-31"
Y, M = 2026, 8


@pytest.fixture()
def seeded():
    reset_all()
    db = database.SessionLocal()
    s = seed_all(db)
    yield s  # держим сессию открытой — объекты остаются привязаны
    db.close()


def _xlsx_cells(content: bytes):
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    out = {ws.title: [list(r) for r in ws.iter_rows(values_only=True)] for ws in wb.worksheets}
    wb.close()
    return out


def _endpoint_fp(client, s, actor) -> dict:
    """Слепок всех endpoint-точек карты ОТ ЛИЦА actor. Endpoints вне прав вернут
    non-200 — для одного и того же actor это одинаково при hook OFF и ON."""
    h = auth_headers(actor)
    fp: dict = {}

    def get(key, url, **params):
        r = client.get(url, params=params or None, headers=h)
        fp[f"GET {key}"] = ("json", r.json()) if r.status_code == 200 else ("err", r.status_code)

    def get_xlsx(key, url, **params):
        r = client.get(url, params=params or None, headers=h)
        fp[f"XLSX {key}"] = (r.status_code, _xlsx_cells(r.content) if r.status_code == 200 else None)

    get("expenses", "/api/expenses")
    get("income", "/api/income")
    get("income_mine", "/api/income/mine")
    get("transfers", "/api/transfers")
    get("income_sources", "/api/income-sources")
    get("departments", "/api/departments")
    get("users", "/api/users")
    get("dashboard", "/api/dashboard")
    get("r_summary", "/api/reports/summary", **{"from": FROM, "to": TO})
    get("r_by_employee", "/api/reports/by-employee", **{"from": FROM, "to": TO})
    get("r_by_category", "/api/reports/by-category", **{"from": FROM, "to": TO})
    get("r_balances", "/api/reports/balances")
    get("r_categories", "/api/reports/categories", year=Y, month=M)
    get("r_employees", "/api/reports/employees", year=Y, month=M)
    get("r_by_department", "/api/reports/by-department", year=Y, month=M)
    get("r_incomes", "/api/reports/incomes", year=Y, month=M)
    get("r_balance", "/api/reports/balance", **{"from": FROM, "to": TO})
    get_xlsx("x_excel", "/api/reports/excel", **{"from": FROM, "to": TO})
    get_xlsx("x_categories", "/api/reports/categories.xlsx", year=Y, month=M)
    get_xlsx("x_employees", "/api/reports/employees.xlsx", year=Y, month=M)
    get_xlsx("x_emp_details", f"/api/reports/employees/{s.acc1.id}/details.xlsx", year=Y, month=M)
    get_xlsx("x_emp_history", f"/api/reports/employees/{s.acc1.id}/history.xlsx")
    get_xlsx("x_expenses_export", "/api/expenses/export.xlsx")
    for u in (s.director, s.acc1, s.acc2, s.conf, s.wsowner, s.auditor):
        get(f"balance_{u.id}", f"/api/users/{u.id}/balance")
        get(f"profile_{u.id}", f"/api/employees/{u.id}/profile", year=Y, month=M)
    get("recent_ops", "/api/admin/recent-operations", limit=100)
    r = client.post("/api/admin/find-duplicates", json={}, headers=h)
    fp["POST find_duplicates"] = ("json", r.json()) if r.status_code == 200 else ("err", r.status_code)
    get("ws_list", "/api/workspaces")
    get("ws_summary", f"/api/workspaces/{s.ws.id}/summary")
    get("ws_member_balances", f"/api/workspaces/{s.ws.id}/members/balances")
    get("ws_by_category", f"/api/workspaces/{s.ws.id}/reports/by-category")
    get("ws_expenses", f"/api/workspaces/{s.ws.id}/expenses")
    return fp


def _balance_fp(db, s) -> dict:
    fp = {}
    for u in (s.director, s.acc1, s.acc2, s.conf, s.wsowner, s.auditor):
        fp[f"cur_{u.id}"] = str(compute_current_balance(db, s.org.id, u.id))
        fp[f"rec_{u.id}"] = str(compute_total_received(db, s.org.id, u.id))
        fp[f"iss_{u.id}"] = str(compute_total_issued(db, s.org.id, u.id))
        fp[f"spent_{u.id}"] = str(spent_total(db, s.org.id, u.id))
        fp[f"pend_{u.id}"] = str(pending_total(db, s.org.id, u.id))
        fp[f"bycur_{u.id}"] = {k: str(v) for k, v in
                               compute_balances_by_currency(db, s.org.id, u.id).items()}
    return fp


def _canon(v):
    """Канонизация: списки сортируем — сравнение OFF/ON независимо от порядка строк
    (при равных таймстампах в seed БД возвращает тай-строки в произвольном порядке,
    а хук слегка меняет план запроса → другой порядок ТЕХ ЖЕ строк, не баг)."""
    if isinstance(v, tuple):  # ('json', {...}) / (200, {...}) — позиция важна, не сортируем
        return tuple(_canon(x) for x in v)
    if isinstance(v, dict):
        return {k: _canon(v[k]) for k in v}
    if isinstance(v, list):
        items = [_canon(x) for x in v]
        return sorted(items, key=lambda x: json.dumps(x, sort_keys=True, ensure_ascii=False, default=str))
    return v


def _diff(a, b):
    return [f"  {k}:\n    OFF={a.get(k)}\n    ON ={b.get(k)}"
            for k in sorted(a.keys() | b.keys()) if _canon(a.get(k)) != _canon(b.get(k))]


ACTOR_NAMES = ["director", "accountable", "auditor", "ws_member", "confidential"]


@pytest.mark.parametrize("actor_name", ACTOR_NAMES)
def test_zero_diff_per_role(client, seeded, actor_name):
    """Прогон 1: на данных без удалённых хук не меняет НИ ОДНОЙ цифры — для каждой роли."""
    s = seeded
    actor = {"director": s.director, "accountable": s.acc1, "auditor": s.auditor,
             "ws_member": s.wsowner, "confidential": s.conf}[actor_name]
    db = database.SessionLocal()
    try:
        disable_hook()
        off = _endpoint_fp(client, s, actor)
        off_bal = _balance_fp(db, s)
        enable_hook()
        db.expire_all()
        on = _endpoint_fp(client, s, actor)
        on_bal = _balance_fp(db, s)
    finally:
        enable_hook()
        db.close()

    if actor_name == "director":
        errs = [k for k, v in on.items() if (k.startswith("GET ") and v[0] == "err")
                or (k.startswith("XLSX ") and v[0] != 200)]
        assert not errs, f"superadmin не получил 200 при включённом хуке: {errs}"

    d = _diff(off, on) + _diff(off_bal, on_bal)
    assert not d, f"[{actor_name}] хук изменил цифры на данных без удалённых:\n" + "\n".join(d)


def _spent(db, s, user):
    return spent_total(db, s.org.id, user.id)


def test_hook_filters_all_record_types(client, seeded):
    """Прогон 2: soft-delete записей РАЗНЫХ типов → они исчезают из агрегатов."""
    s = seeded
    from services.soft_delete import soft_delete
    db = database.SessionLocal()
    try:
        enable_hook()
        h = auth_headers(s.director)

        # снимки ДО
        spent_acc1_before = _spent(db, s, s.acc1)
        spent_acc2_before = _spent(db, s, s.acc2)
        dep_before = client.get("/api/reports/by-department", params={"year": Y, "month": M}, headers=h).json()
        ws_before = client.get(f"/api/workspaces/{s.ws.id}/summary", headers=h).json()

        # soft-delete 5 типов: личный вклад, депозит-покупка, передача-expense, USD, ws-запись
        targets = [s.e_pers, s.e_supp, s.e_transfer, s.e_usd, s.e_ws]
        for obj in targets:
            fresh = db.get(type(obj), obj.id)
            soft_delete(fresh, s.director)
        db.commit()
        db.expire_all()

        # все пятеро исчезли из общего списка расходов
        exp_ids = {e["id"] for e in client.get("/api/expenses", headers=h).json()}
        for obj in targets:
            assert obj.id not in exp_ids, f"{obj.id} остался в списке расходов"

        # USD-запись (e_usd, acc2, 200 USD × 89): spent acc2 упал ровно на 17800 KGS
        assert spent_acc2_before - _spent(db, s, s.acc2) == Decimal("200") * Decimal("89")

        # личный вклад (e_pers, acc1, 2000): spent acc1 упал на 2000
        #   (+ e_transfer тоже acc1 но expense_type=transfer → в spent_total входит? spent_total
        #    считает все статусы pending/approved без фильтра типа → входит. Учитываем обе.)
        #   поэтому проверяем, что личный-вклад ушёл из by-department dept A.
        dep_after = client.get("/api/reports/by-department", params={"year": Y, "month": M}, headers=h).json()
        assert dep_after != dep_before, "by-department не изменился после удаления записей подразделения"

        # ws-запись (e_ws) ушла из пространства
        ws_after = client.get(f"/api/workspaces/{s.ws.id}/summary", headers=h).json()
        assert ws_after != ws_before, "summary пространства не изменилось после удаления его расхода"
        ws_exp_ids = {e["id"] for e in client.get(f"/api/workspaces/{s.ws.id}/expenses", headers=h).json()}
        assert s.e_ws.id not in ws_exp_ids

        # spent acc1 уменьшился хотя бы на личный вклад 2000 (e_pers) + transfer 1000 + supplier 4000
        assert spent_acc1_before - _spent(db, s, s.acc1) == Decimal("2000") + Decimal("1000") + Decimal("4000")
    finally:
        db.close()


@pytest.mark.parametrize("url", [
    "/api/reports/categories", "/api/reports/by-department", "/api/reports/employees",
])
def test_report_timing(client, seeded, url):
    """Тайминг ON/OFF — печать (НЕ доказательство: на seed-объёме план не как на проде)."""
    h = auth_headers(seeded.director)
    params = {"year": Y, "month": M}

    def timeit(n=20):
        t0 = time.perf_counter()
        for _ in range(n):
            client.get(url, params=params, headers=h)
        return (time.perf_counter() - t0) / n * 1000

    disable_hook(); off = timeit()
    enable_hook(); on = timeit()
    print(f"\n[timing] {url}: OFF={off:.1f}ms ON={on:.1f}ms Δ={on-off:+.1f}ms")

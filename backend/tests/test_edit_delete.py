"""Функциональные тесты правки/удаления записей (ТЗ, сценарии 1-14 + доп.).
Каждый тест — чистый мини-набор (org/director/acc/other/depA/depB/cat/курсы).
Маячки: глобальный баланс vs личный вклад; живой курс — оба ДОЛЖНЫ сломаться при
соответствующих будущих правках (B / замороженный курс).
"""
from decimal import Decimal
from types import SimpleNamespace

import pytest

import database
from conftest import auth_headers, reset_all
from models import (
    BalanceTopUp, Category, Department, ExchangeRate, Income,
    Organization, RecordChangeLog, User,
)

D = lambda x: Decimal(str(x))


@pytest.fixture()
def m():
    reset_all()
    db = database.SessionLocal()
    org = Organization(name="Org", plan="legacy"); db.add(org); db.flush()

    def U(name, phone, role="accountable", sup=None):
        u = User(org_id=org.id, name=name, phone=phone, password_hash="x", role=role, supervisor_id=sup)
        db.add(u); db.flush(); return u

    director = U("Дир", "+70000000101", role="superadmin")
    acc = U("Сотрудник", "+70000000102")
    other = U("Подчинённый", "+70000000103", sup=None)
    depA = Department(org_id=org.id, name="A"); depB = Department(org_id=org.id, name="B")
    db.add_all([depA, depB]); db.flush()
    other.supervisor_id = acc.id  # чтобы acc мог делать передачу-расход подчинённому
    cat = Category(org_id=org.id, name="Cat"); db.add(cat)
    db.add_all([
        ExchangeRate(org_id=org.id, from_currency="USD", to_currency="KGS", rate=D("89")),
        ExchangeRate(org_id=org.id, from_currency="RUB", to_currency="KGS", rate=D("0.9")),
    ])
    db.commit()
    ns = SimpleNamespace(org=org, director=director, acc=acc, other=other,
                         depA=depA, depB=depB, cat=cat)
    yield ns
    db.close()


# ---- helpers через реальные эндпоинты ----

def mk_income(client, actor, receiver, amount, currency="KGS"):
    r = client.post("/api/income", json={"amount": str(amount), "currency": currency,
                    "source": "src", "received_by_id": receiver.id}, headers=auth_headers(actor))
    assert r.status_code == 201, r.text
    return r.json()


def mk_expense(client, actor, dep, amount, currency="KGS", cat=None, personal=False,
               to_user=None, spent_at=None):
    body = {"amount": str(amount), "currency": currency, "department_id": dep.id,
            "is_personal_contribution": personal}
    if cat is not None:
        body["category_id"] = cat.id
    if to_user is not None:
        body["to_user_id"] = to_user.id
    if spent_at is not None:
        body["spent_at"] = spent_at
    r = client.post("/api/expenses", json=body, headers=auth_headers(actor))
    assert r.status_code == 201, r.text
    return r.json()


def bal(client, actor, uid):
    r = client.get(f"/api/users/{uid}/balance", headers=auth_headers(actor))
    assert r.status_code == 200, r.text
    return D(r.json()["current_balance"])


def review(client, director, eid, status, comment=None):
    r = client.post(f"/api/expenses/{eid}/review", json={"status": status, "review_comment": comment},
                    headers=auth_headers(director))
    assert r.status_code == 200, r.text


def dep_summary(client, director, dep_id):
    r = client.get("/api/reports/by-department", params={"year": 2026, "month": 8},
                   headers=auth_headers(director))
    for d in r.json()["departments"]:
        if d["id"] == dep_id:
            return d
    return {"summary": {"received": 0.0, "spent": 0.0}, "employees": []}


def last_log(entity_type, entity_id):
    db = database.SessionLocal()
    try:
        return (db.query(RecordChangeLog)
                .filter_by(entity_type=entity_type, entity_id=entity_id)
                .order_by(RecordChangeLog.id.desc()).first())
    finally:
        db.close()


# ===================== ТЗ 1-14 =====================

def test_01_delete_expense_restores_balance(client, m):
    mk_income(client, m.acc, m.acc, 10000)
    e = mk_expense(client, m.acc, m.depA, 3000)
    assert bal(client, m.director, m.acc.id) == D(7000)
    assert client.delete(f"/api/expenses/{e['id']}", headers=auth_headers(m.acc)).status_code == 204
    assert bal(client, m.director, m.acc.id) == D(10000)


def test_02_edit_amount_recomputes_balance(client, m):
    mk_income(client, m.acc, m.acc, 10000)
    e = mk_expense(client, m.acc, m.depA, 3000)
    r = client.patch(f"/api/expenses/{e['id']}", json={"amount": "4500"}, headers=auth_headers(m.acc))
    assert r.status_code == 200, r.text
    assert bal(client, m.director, m.acc.id) == D(5500)


def test_03_personal_contribution_department_invariant(client, m):
    """Личный вклад в разрезе подразделения = +приход и +расход; удаление откатывает оба."""
    base = dep_summary(client, m.director, m.depA.id)["summary"]
    e = mk_expense(client, m.acc, m.depA, 2000, cat=m.cat, personal=True)
    after = dep_summary(client, m.director, m.depA.id)["summary"]
    assert after["spent"] == base["spent"] + 2000
    assert after["received"] == base["received"] + 2000   # тот же расход зачтён приходом
    # удаление откатывает обе величины
    assert client.delete(f"/api/expenses/{e['id']}", headers=auth_headers(m.acc)).status_code == 204
    back = dep_summary(client, m.director, m.depA.id)["summary"]
    assert back["spent"] == base["spent"]
    assert back["received"] == base["received"]


def test_04_toggle_personal_contribution_both_ways(client, m):
    e = mk_expense(client, m.acc, m.depA, 2000, cat=m.cat, personal=False)
    base = dep_summary(client, m.director, m.depA.id)["summary"]
    # false → true: приход подразделения +2000
    r = client.post(f"/api/expenses/{e['id']}/personal-contribution", json={"enabled": True},
                    headers=auth_headers(m.acc))
    assert r.status_code == 200
    on = dep_summary(client, m.director, m.depA.id)["summary"]
    assert on["received"] == base["received"] + 2000
    # true → false: приход возвращается
    r = client.post(f"/api/expenses/{e['id']}/personal-contribution", json={"enabled": False},
                    headers=auth_headers(m.acc))
    assert r.status_code == 200
    off = dep_summary(client, m.director, m.depA.id)["summary"]
    assert off["received"] == base["received"]


def test_05_change_department_moves_sum(client, m):
    e = mk_expense(client, m.acc, m.depA, 3000, cat=m.cat)
    a0 = dep_summary(client, m.director, m.depA.id)["summary"]["spent"]
    b0 = dep_summary(client, m.director, m.depB.id)["summary"]["spent"]
    r = client.patch(f"/api/expenses/{e['id']}", json={"department_id": m.depB.id},
                     headers=auth_headers(m.acc))
    assert r.status_code == 200, r.text
    a1 = dep_summary(client, m.director, m.depA.id)["summary"]["spent"]
    b1 = dep_summary(client, m.director, m.depB.id)["summary"]["spent"]
    assert a0 - a1 == 3000 and b1 - b0 == 3000


def test_06_change_currency_recomputes_kgs_neighbor_untouched(client, m):
    neighbor = mk_expense(client, m.acc, m.depA, 500, currency="KGS")
    e = mk_expense(client, m.acc, m.depA, 1000, currency="RUB")   # amount_kgs = 1000×0.9 = 900
    assert D(e["amount_kgs"]) == D(900)
    r = client.patch(f"/api/expenses/{e['id']}", json={"currency": "KGS"}, headers=auth_headers(m.acc))
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["currency"] == "KGS" and D(body["amount"]) == D(1000) and D(body["amount_kgs"]) == D(1000)
    # сосед не затронут
    n = client.get(f"/api/expenses/{neighbor['id']}", headers=auth_headers(m.acc)).json()
    assert D(n["amount_kgs"]) == D(500)


def test_07_change_date_between_periods(client, m):
    e = mk_expense(client, m.acc, m.depA, 3000, spent_at="2026-07-31T10:00:00")

    def in_range(frm, to):
        r = client.get("/api/expenses", params={"date_from": frm, "date_to": to}, headers=auth_headers(m.acc))
        return e["id"] in [x["id"] for x in r.json()]

    assert in_range("2026-07-01", "2026-08-01") and not in_range("2026-08-01", "2026-09-01")
    r = client.patch(f"/api/expenses/{e['id']}", json={"spent_at": "2026-08-01T10:00:00"},
                     headers=auth_headers(m.acc))
    assert r.status_code == 200
    assert not in_range("2026-07-01", "2026-08-01") and in_range("2026-08-01", "2026-09-01")


def test_08_delete_transfer_reverts_both(client, m):
    mk_income(client, m.acc, m.acc, 10000)
    r = client.post("/api/transfers", json={"to_user_id": m.other.id, "amount": "5000", "currency": "KGS"},
                    headers=auth_headers(m.acc))
    assert r.status_code == 201, r.text
    tid = r.json()["id"]
    assert bal(client, m.director, m.acc.id) == D(5000)
    assert bal(client, m.director, m.other.id) == D(5000)
    assert client.delete(f"/api/transfers/{tid}", headers=auth_headers(m.acc)).status_code == 204
    assert bal(client, m.director, m.acc.id) == D(10000)
    assert bal(client, m.director, m.other.id) == D(0)


def test_09_employee_cannot_touch_approved(client, m):
    e = mk_expense(client, m.acc, m.depA, 3000)
    review(client, m.director, e["id"], "approved")
    assert client.patch(f"/api/expenses/{e['id']}", json={"amount": "1"}, headers=auth_headers(m.acc)).status_code == 403
    assert client.delete(f"/api/expenses/{e['id']}", headers=auth_headers(m.acc)).status_code == 403
    got = client.get(f"/api/expenses/{e['id']}", headers=auth_headers(m.director)).json()
    assert D(got["amount"]) == D(3000) and got["status"] == "approved"


def test_10_cannot_touch_others_record(client, m):
    e = mk_expense(client, m.acc, m.depA, 3000)
    assert client.patch(f"/api/expenses/{e['id']}", json={"amount": "1"}, headers=auth_headers(m.other)).status_code == 403
    assert client.delete(f"/api/expenses/{e['id']}", headers=auth_headers(m.other)).status_code == 403


def test_11_edit_rejected_returns_to_pending(client, m):
    e = mk_expense(client, m.acc, m.depA, 3000)
    review(client, m.director, e["id"], "rejected", "нет чека")
    r = client.patch(f"/api/expenses/{e['id']}", json={"amount": "2500"}, headers=auth_headers(m.acc))
    assert r.status_code == 200, r.text
    assert r.json()["status"] == "pending"


def test_12_double_delete_404_no_double_effect(client, m):
    mk_income(client, m.acc, m.acc, 10000)
    e = mk_expense(client, m.acc, m.depA, 3000)
    assert client.delete(f"/api/expenses/{e['id']}", headers=auth_headers(m.acc)).status_code == 204
    assert bal(client, m.director, m.acc.id) == D(10000)
    assert client.delete(f"/api/expenses/{e['id']}", headers=auth_headers(m.acc)).status_code == 404
    assert bal(client, m.director, m.acc.id) == D(10000)   # не вычлось второй раз


def test_13_deleted_absent_from_list_and_excel(client, m):
    e = mk_expense(client, m.acc, m.depA, 3000)
    client.delete(f"/api/expenses/{e['id']}", headers=auth_headers(m.acc))
    ids = [x["id"] for x in client.get("/api/expenses", headers=auth_headers(m.director)).json()]
    assert e["id"] not in ids
    # Excel-экспорт не содержит удалённой строки
    import io
    from openpyxl import load_workbook
    r = client.get("/api/expenses/export.xlsx", headers=auth_headers(m.director))
    wb = load_workbook(io.BytesIO(r.content), read_only=True)
    rows = [list(row) for row in wb.active.iter_rows(values_only=True)]
    wb.close()
    assert all(not (len(row) > 3 and row[3] == 3000) for row in rows[1:])


def test_14_change_log_written(client, m):
    e = mk_expense(client, m.acc, m.depA, 3000)
    client.patch(f"/api/expenses/{e['id']}", json={"amount": "4500"}, headers=auth_headers(m.acc))
    up = last_log("expense", e["id"])
    assert up is not None and up.action == "update"
    amt = up.diff.get("amount")   # значения — валидные Decimal-строки (для восстановимости)
    assert D(amt["old"]) == D(3000) and D(amt["new"]) == D(4500)
    client.delete(f"/api/expenses/{e['id']}", headers=auth_headers(m.acc))
    dl = last_log("expense", e["id"])
    assert dl.action == "delete" and D(dl.diff.get("amount")) == D(4500)  # полный снимок записи


# ===================== доп. кейсы =====================

def test_income_created_by_other_forbidden_for_employee(client, m):
    inc = mk_income(client, m.director, m.acc, 5000)  # внёс директор, получатель acc
    assert client.delete(f"/api/income/{inc['id']}", headers=auth_headers(m.acc)).status_code == 403
    assert client.patch(f"/api/income/{inc['id']}", json={"amount": "1"}, headers=auth_headers(m.acc)).status_code == 403


def test_income_own_editable_by_employee(client, m):
    inc = mk_income(client, m.acc, m.acc, 5000)  # свой ручной
    assert client.patch(f"/api/income/{inc['id']}", json={"amount": "6000"}, headers=auth_headers(m.acc)).status_code == 200
    assert client.delete(f"/api/income/{inc['id']}", headers=auth_headers(m.acc)).status_code == 204


def test_delete_transfer_expense_removes_paired_topup(client, m):
    """Удаление расхода-передачи снимает парный BalanceTopUp (чиним осиротение)."""
    mk_income(client, m.acc, m.acc, 10000)
    e = mk_expense(client, m.acc, m.depA, 5000, to_user=m.other)  # transfer-expense + парный topup
    assert bal(client, m.director, m.other.id) == D(5000)   # получатель получил
    assert client.delete(f"/api/expenses/{e['id']}", headers=auth_headers(m.acc)).status_code == 204
    assert bal(client, m.director, m.other.id) == D(0)      # парный topup тоже ушёл


# ===================== МАЯЧКИ (должны сломаться при будущих правках) =====================

def test_beacon_personal_contribution_currently_lowers_global_balance(client, m):
    """МАЯЧОК: сейчас глобальный current_balance УМЕНЬШАЕТСЯ на личный вклад, что
    ПРОТИВОРЕЧИТ тексту под чекбоксом («личный баланс не меняется»). Это баг, подлежит
    исправлению (задача B). Тест ДОЛЖЕН СЛОМАТЬСЯ, когда B исключит is_personal_contribution
    из _expenses_approved — тогда поставим == b0 и уберём маячок."""
    mk_income(client, m.acc, m.acc, 10000)
    b0 = bal(client, m.director, m.acc.id)
    mk_expense(client, m.acc, m.depA, 2000, cat=m.cat, personal=True)
    b1 = bal(client, m.director, m.acc.id)
    assert b1 == b0 - D(2000), (
        "МАЯЧОК СРАБОТАЛ: личный вклад больше не уменьшает глобальный баланс — "
        "значит сделана задача B. Обнови тест 3/удали маячок."
    )


def test_dashboard_native_currency_single(client, m):
    """(B) Сотрудник с активностью в ОДНОЙ валюте → дашборд в этой валюте (не сом)."""
    mk_expense(client, m.acc, m.depA, 1000, currency="RUB")  # amount_kgs = 900 (курс 0.9)
    t = client.get("/api/dashboard", headers=auth_headers(m.acc)).json()["totals"]
    assert t["balance_currency"] == "RUB"
    assert abs(t["spent"] - 1000) < 0.01          # родные рубли, не 900 сом
    assert abs(t["current_balance"] + 1000) < 0.01  # −1000 ₽


def test_dashboard_som_when_mixed(client, m):
    """(B) Смешанные валюты → сом-эквивалент (KGS)."""
    mk_expense(client, m.acc, m.depA, 1000, currency="RUB")
    mk_expense(client, m.acc, m.depA, 100, currency="USD")
    t = client.get("/api/dashboard", headers=auth_headers(m.acc)).json()["totals"]
    assert t["balance_currency"] == "KGS"


def test_beacon_live_rate_changes_historical_kgs(client, m):
    """МАЯЧОК: баланс считается по ЖИВОМУ курсу — при смене курса исторические
    KGS-суммы меняются задним числом (осознанное действующее поведение). Тест ДОЛЖЕН
    СЛОМАТЬСЯ при переходе на замороженный per-record курс (fx_rate/fx_rate_date)."""
    mk_expense(client, m.acc, m.depA, 100, currency="USD")   # 100×89 = 8900 KGS
    b0 = bal(client, m.director, m.acc.id)
    db = database.SessionLocal()
    db.add(ExchangeRate(org_id=m.org.id, from_currency="USD", to_currency="KGS", rate=D("100")))
    db.commit(); db.close()
    b1 = bal(client, m.director, m.acc.id)   # теперь 100×100 = 10000 → баланс сдвинулся
    assert b1 != b0, (
        "МАЯЧОК СРАБОТАЛ: смена курса больше не двигает исторические суммы — "
        "значит введён замороженный курс. Пересмотри тест 6."
    )
